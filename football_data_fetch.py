import os
import requests
import openpyxl
from collections import defaultdict
from datetime import datetime, timedelta, time, timezone
from pathlib import Path
import json
import re
from concurrent.futures import ThreadPoolExecutor, as_completed

# Slack: token from env SLACK_BOT_TOKEN or from .slack_bot_token (one line, same folder as script)
SLACK_CHANNEL = "#notifications-fmm"
SLACK_TOKEN_FILE = Path(__file__).resolve().parent / ".slack_bot_token"

# Leagues where the API doesn't return an abbreviation – we use these for HS/webcast URLs
LEAGUE_ABBREV_OVERRIDE = {
    61: "MFL",
    13: "LFF",
    31: "CNCC",
    101: "CRC",
    56: "VPF",
    64: "LDF",
    37: "ESFL",
    43: "LNFPH",
    91: "MNL",
    77: "IFA",
    58: "GOAFA",
    95: "JPL",
    79: "FPNMF",
    42: "CTFA",
    72: "COS",
}

# Caches so we don't hit the API twice for the same league or match
_league_details_cache = {}
_match_details_cache = {}

def get_user_api_key():
    """Prompt for the Genius Sports API key."""
    print("=== FOOTBALL DATA FETCH ===")
    api_key = input("Enter your API key: ").strip()
    if not api_key:
        print("API key is required!")
        return None
    return api_key


def _get_slack_token():
    """Slack bot token: env SLACK_BOT_TOKEN or first line of .slack_bot_token in script folder."""
    token = os.environ.get("SLACK_BOT_TOKEN", "").strip()
    if token:
        return token
    if SLACK_TOKEN_FILE.exists():
        try:
            return SLACK_TOKEN_FILE.read_text(encoding="utf-8").strip().splitlines()[0].strip()
        except Exception:
            pass
    return ""


def _match_has_started(match):
    """True if kickoff is in the past (game started). Prefers matchTimeUTC when present."""
    utc_str = match.get("matchTimeUTC", "").strip()
    if utc_str:
        try:
            kickoff = datetime.strptime(utc_str, "%Y-%m-%d %H:%M:%S").replace(tzinfo=timezone.utc)
            return kickoff <= datetime.now(timezone.utc)
        except Exception:
            pass
    local_str = match.get("matchTime", "").strip()
    if local_str:
        try:
            kickoff = datetime.strptime(local_str, "%Y-%m-%d %H:%M:%S")
            return kickoff <= datetime.now()
        except Exception:
            pass
    return False


def collect_check_issues(competitions):
    """
    Build a list of matches that failed one or more checks (DM, WHST, Statistician, Webcast, End game).
    Used for the Slack issues section. Returns per-competition blocks with match details and webcast/HS links.
    """
    result = []
    for comp in competitions:
        league_name = comp.get("leagueName", "")
        league_id = comp.get("leagueId", "")
        competition_name = comp.get("competitionName", "")
        competition_id = comp.get("competitionId", "")
        problem_matches = []
        for match in comp.get("matches", []):
            failed = []
            webcast_url = None
            hs_url = None
            if (match.get("livestream_status") or "").strip() == "No":
                failed.append("Pre-game DM check")
            if (match.get("whst_live_data_source_match") or "").strip() == "No":
                failed.append("Pre-game WHST Live Data Source")
            started = _match_has_started(match)
            if started:
                if (match.get("publish_connection_status") or "").strip() == "No":
                    failed.append("Live game Statistician check")
                status = (match.get("webcast_status") or "").strip()
                if status in ("No", "N/A"):
                    failed.append("Live game Webcast check")
                    mid = match.get("matchId", "")
                    league_abbrev = (comp.get("leagueAbbrev") or "").strip().lower()
                    if mid and league_abbrev:
                        webcast_url = f"https://livestats.dcd.shared.geniussports.com/u/{league_abbrev}/{mid}/"
            end_game = (match.get("end_game_status") or "").strip()
            if end_game == "Match check required":
                failed.append("End game past match data")
                hs_url = match.get("end_game_hs_url") or ""
            elif end_game == "N/A - Match check required":
                failed.append("End game past match data")
                hs_url = match.get("end_game_hs_url") or ""
            if failed:
                problem_matches.append({
                    "game": (match.get("game") or "").strip() or "Match",
                    "matchId": match.get("matchId", ""),
                    "failed_checks": failed,
                    "webcast_url": webcast_url,
                    "hs_url": hs_url,
                })
        if problem_matches:
            result.append({
                "league_name": league_name,
                "league_id": league_id,
                "competition_name": competition_name,
                "competition_id": competition_id,
                "matches": problem_matches,
            })
    return result


def send_slack_message(text, channel=SLACK_CHANNEL, blocks=None):
    """Post message to Slack. Returns True if sent, False if no token or send failed."""
    token = _get_slack_token()
    if not token:
        return False
    try:
        from slack_sdk import WebClient
        from slack_sdk.errors import SlackApiError
        client = WebClient(token=token)
        kwargs = {"channel": channel, "text": text}
        if blocks:
            kwargs["blocks"] = blocks
        client.chat_postMessage(**kwargs)
        return True
    except Exception as e:
        print(f"⚠️  Slack notification failed: {e}")
        return False

def fetch_all_leagues(api_key):
    """Fetch all leagues from the Genius Sports API (used when not using whitelist-only flow)."""
    print("\n=== FETCHING ALL LEAGUES ===")
    
    url = "https://api.wh.geniussports.com/v1/football/leagues"
    params = {
        'ak': api_key,
        'limit': 500
    }
    
    try:
        print(f"Making API call to: {url}")
        print(f"Parameters: {params}")
        
        response = requests.get(url, params=params, timeout=30)
        print(f"Response status code: {response.status_code}")
        
        if response.status_code != 200:
            print(f"API returned error status: {response.status_code}")
            print(f"Response content: {response.text[:500]}...")
            return None
        
        data = response.json()
        print(f"API call successful. Response keys: {list(data.keys())}")
        
        # Extract leagues from response
        leagues = []
        if 'response' in data and 'data' in data['response']:
            leagues_data = data['response']['data']
            print(f"Found {len(leagues_data)} leagues")
            
            for league in leagues_data:
                league_info = {
                    'leagueId': league.get('leagueId', ''),
                    'leagueName': league.get('leagueName', ''),
                    'country': league.get('country', ''),
                    'countryCode': league.get('countryCode', ''),
                    'timezone': league.get('timezone', ''),
                    'competitionStandard': league.get('competitionStandard', '')
                }
                leagues.append(league_info)
                print(f"  - {league_info['leagueName']} (ID: {league_info['leagueId']}) - {league_info['country']}")
        
        return leagues
        
    except requests.exceptions.RequestException as e:
        print(f"API call failed: {str(e)}")
        return None
    except Exception as e:
        print(f"Error processing API response: {str(e)}")
        return None

def load_competition_whitelist(whitelist_file='competition_whitelist.json'):
    """Load which competitions to process from competition_whitelist.json."""
    try:
        with open(whitelist_file, 'r', encoding='utf-8') as f:
            config = json.load(f)
        
        # Extract competition IDs from the whitelist
        competition_ids = [comp['id'] for comp in config['active_competitions']]
        
        print(f"Loaded whitelist with {len(competition_ids)} competitions:")
        for comp in config['active_competitions']:
            print(f"  - {comp['name']} (ID: {comp['id']})")
        
        return competition_ids, config
        
    except FileNotFoundError:
        print(f"Whitelist file not found: {whitelist_file}")
        print("Creating empty whitelist. Please add competitions manually to the JSON file.")
        return [], {}
    except json.JSONDecodeError as e:
        print(f"Error parsing whitelist file: {e}")
        return [], {}
    except Exception as e:
        print(f"Error loading whitelist: {e}")
        return [], {}

def filter_competitions_by_whitelist(competitions, whitelist_ids):
    """Keep only competitions whose IDs are in the whitelist."""
    print(f"\n=== FILTERING COMPETITIONS BY WHITELIST ===")
    print(f"Total competitions found: {len(competitions)}")
    print(f"Whitelist contains: {len(whitelist_ids)} competition IDs")
    
    filtered_competitions = []
    
    for comp in competitions:
        if comp['competitionId'] in whitelist_ids:
            filtered_competitions.append(comp)
            print(f"  ✓ {comp['competitionName']} (ID: {comp['competitionId']}) - WHITELISTED")
        else:
            print(f"  ✗ {comp['competitionName']} (ID: {comp['competitionId']}) - NOT in whitelist")
    
    print(f"Filtered to {len(filtered_competitions)} whitelisted competitions")
    return filtered_competitions

def fetch_league_details(api_key, league_id):
    """League abbreviation for HS/webcast URLs. Uses overrides first, then cache, then API."""
    if not league_id:
        return None
    
    cache_key = f"{league_id}"
    if cache_key in _league_details_cache:
        return _league_details_cache[cache_key]
    
    try:
        lid = int(league_id)
        if lid in LEAGUE_ABBREV_OVERRIDE:
            abbr = LEAGUE_ABBREV_OVERRIDE[lid]
            _league_details_cache[cache_key] = abbr
            return abbr
    except (TypeError, ValueError):
        pass
    
    try:
        url = f"https://api.wh.geniussports.com/v1/football/leagues/{league_id}"
        params = { 'ak': api_key }
        resp = requests.get(url, params=params, timeout=30)
        
        if resp.status_code == 200:
            data = resp.json()
            league_data = data.get('response', {}).get('data') or data.get('response') or data
            if not isinstance(league_data, dict):
                league_data = {}
            league_abbrev = (
                league_data.get('leagueAbbrev') or league_data.get('abbreviation') or
                league_data.get('federationCode') or league_data.get('code') or league_data.get('slug')
            )
            if not league_abbrev:
                try:
                    lid = int(league_id)
                    if lid in LEAGUE_ABBREV_OVERRIDE:
                        league_abbrev = LEAGUE_ABBREV_OVERRIDE[lid]
                except (TypeError, ValueError):
                    pass
            _league_details_cache[cache_key] = league_abbrev
            return league_abbrev
        else:
            print(f"  ⚠️  Could not fetch league details (HTTP {resp.status_code})")
            try:
                lid = int(league_id)
                if lid in LEAGUE_ABBREV_OVERRIDE:
                    abbr = LEAGUE_ABBREV_OVERRIDE[lid]
                    _league_details_cache[cache_key] = abbr
                    return abbr
            except (TypeError, ValueError):
                pass
            return None
    except Exception as e:
        print(f"  ⚠️  Error fetching league details: {e}")
        try:
            lid = int(league_id)
            if lid in LEAGUE_ABBREV_OVERRIDE:
                abbr = LEAGUE_ABBREV_OVERRIDE[lid]
                _league_details_cache[cache_key] = abbr
                return abbr
        except (TypeError, ValueError):
            pass
        return None

def process_whitelisted_competitions_directly(api_key, whitelist_ids, whitelist_config):
    """Load league abbrev and LiveDataSource per whitelist entry, then fetch matches for each competition."""
    competitions_with_matches = []
    
    # Get whitelist data for names and IDs
    whitelist_competitions = whitelist_config.get('active_competitions', [])
    whitelist_lookup = {comp['id']: comp for comp in whitelist_competitions}
    
    for comp_id in whitelist_ids:
        print(f"\nProcessing competition ID: {comp_id}")
        
        comp_data = whitelist_lookup.get(comp_id, {})
        comp_name = comp_data.get('name', f"Competition {comp_id}")
        league_id = comp_data.get('league_id', 0)
        league_name = comp_data.get('league_name', "Unknown League")
        live_data_source = None
        league_abbrev = None
        
        if league_id:
            league_abbrev = fetch_league_details(api_key, league_id)
            if not league_abbrev and comp_data.get('federation_code'):
                league_abbrev = str(comp_data.get('federation_code')).strip()
                print(f"  League Abbreviation: {league_abbrev} (from whitelist)")
            elif league_abbrev:
                print(f"  League Abbreviation: {league_abbrev}")
        
        if league_id:
            try:
                url = f"https://api.wh.geniussports.com/v1/football/leagues/{league_id}/competitions"
                params = { 'ak': api_key, 'limit': 100 }
                resp = requests.get(url, params=params, timeout=30)
                if resp.status_code == 200:
                    comp_list = resp.json().get('response', {}).get('data', [])
                    for c in comp_list:
                        if c.get('competitionId') == comp_id:
                            live_data_source = (
                                (c.get('internalConfiguration') or {}).get('LiveDataSource')
                                or c.get('competitionLiveDataSource')
                            )
                            break
                else:
                    print(f"  ⚠️  Could not fetch competition details for LiveDataSource (HTTP {resp.status_code})")
            except Exception as e:
                print(f"  ⚠️  Error fetching LiveDataSource: {e}")
        
        print(f"  Competition: {comp_name}")
        print(f"  League: {league_name} (ID: {league_id})")
        if live_data_source:
            print(f"  LiveDataSource: {live_data_source}")
        
        matches = fetch_matches_for_competition(api_key, comp_id, comp_name, live_data_source, league_abbrev)
        
        # Create competition entry (always include, even if no matches)
        competition_entry = {
            'competitionId': comp_id,
            'competitionName': comp_name,
            'leagueId': league_id,
            'leagueName': league_name,
            'leagueAbbrev': league_abbrev or '',
            'liveDataSource': live_data_source or '',
            'matches': matches if matches else []
        }
        
        competitions_with_matches.append(competition_entry)
        
        if matches:
            print(f"  ✅ Found {len(matches)} matches for competition {comp_id}")
        else:
            print(f"  ⚠️  No matches found for competition {comp_id} (table will be empty)")
    
    return competitions_with_matches

def show_whitelist_management_info(whitelist_config):
    """
    Show information about managing the whitelist
    """
    print(f"\n=== WHITELIST MANAGEMENT INFO ===")
    print(f"Last updated: {whitelist_config.get('last_updated', 'Unknown')}")
    print(f"Total competitions in whitelist: {whitelist_config.get('total_competitions', 0)}")
    
    if whitelist_config.get('configuration', {}).get('require_manual_approval', True):
        print("\nTo add new competitions:")
        print("1. Find the competition ID from the API or your Confluence list")
        print("2. Edit competition_whitelist.json")
        print("3. Add a new object to the 'active_competitions' array")
        print("4. Run the script again")
    
    print(f"\nExample JSON entry:")
    print('''{
    "id": 1234,
    "name": "Competition Name",
    "league_id": 567,
    "league_name": "League Name",
    "priority": "high",
    "notes": "Added from Slack message"
}''')

def fetch_competitions_for_league(api_key, league_id, league_name):
    """
    Fetch competitions for a specific league
    """
    print(f"  Fetching competitions for league: {league_name} (ID: {league_id})")
    
    url = f"https://api.wh.geniussports.com/v1/football/leagues/{league_id}/competitions"
    params = {
        'ak': api_key,
        'limit': 100
    }
    
    try:
        response = requests.get(url, params=params, timeout=30)
        
        if response.status_code != 200:
            print(f"    API returned error status: {response.status_code}")
            return []
        
        data = response.json()
        
        # Extract competitions from response
        competitions = []
        if 'response' in data and 'data' in data['response']:
            competitions_data = data['response']['data']
            print(f"    Found {len(competitions_data)} competitions")
            
            for competition in competitions_data:
                competition_info = {
                    'competitionId': competition.get('competitionId', ''),
                    'competitionName': competition.get('competitionNameInternational', '') or competition.get('competitionName', ''),
                    'startDate': competition.get('startDate', ''),
                    'endDate': competition.get('endDate', ''),
                    'leagueId': league_id,
                    'leagueName': league_name
                }
                competitions.append(competition_info)
                print(f"      - {competition_info['competitionName']} (ID: {competition_info['competitionId']}) - {competition_info['startDate']} to {competition_info['endDate']}")
        
        return competitions
        
    except requests.exceptions.RequestException as e:
        print(f"    API call failed: {str(e)}")
        return []
    except Exception as e:
        print(f"    Error processing API response: {str(e)}")
        return []


def fetch_all_competitions(api_key, leagues):
    """
    Fetch competitions for all leagues
    """
    print(f"\n=== FETCHING COMPETITIONS FOR ALL LEAGUES ===")
    
    all_competitions = []
    
    for league_idx, league in enumerate(leagues):
        print(f"\nProcessing league {league_idx + 1}/{len(leagues)}: {league['leagueName']}")
        
        competitions = fetch_competitions_for_league(api_key, league['leagueId'], league['leagueName'])
        all_competitions.extend(competitions)
    
    print(f"\nTotal competitions found: {len(all_competitions)}")
    
    # Filter competitions by date
    filtered_competitions = filter_competitions_by_date(all_competitions)
    
    return filtered_competitions

def filter_matches_by_date(matches):
    """
    Filter matches to include only those from today to 1 week ahead
    """
    today = datetime.now().date()
    one_week_ahead = today + timedelta(days=7)
    
    print(f"    Date filter: today={today}, one_week_ahead={one_week_ahead}")
    
    filtered_matches = []
    
    for match in matches:
        match_time_str = match.get('matchTime', '')
        
        if not match_time_str:
            print(f"    Skipping match with no matchTime: {match.get('game', 'Unknown')}")
            continue
        
        try:
            print(f"    Raw matchTime string: '{match_time_str}'")
            
            # Try different date formats that might be returned by the API
            match_datetime = None
            date_formats = [
                '%Y-%m-%d %H:%M:%S',  # 2024-10-16 14:30:00
                '%Y-%m-%d %H:%M',     # 2024-10-16 14:30
                '%Y-%m-%d',           # 2024-10-16
                '%d/%m/%Y %H:%M:%S',  # 16/10/2024 14:30:00
                '%d/%m/%Y %H:%M',     # 16/10/2024 14:30
                '%d/%m/%Y',           # 16/10/2024
                '%m/%d/%Y %H:%M:%S',  # 10/16/2024 14:30:00 (US format)
                '%m/%d/%Y %H:%M',     # 10/16/2024 14:30
                '%m/%d/%Y',           # 10/16/2024
            ]
            
            for date_format in date_formats:
                try:
                    match_datetime = datetime.strptime(match_time_str, date_format)
                    print(f"    Successfully parsed with format: {date_format}")
                    break
                except ValueError:
                    continue
            
            if match_datetime is None:
                raise ValueError(f"Could not parse date with any known format")
            
            match_date = match_datetime.date()
            
            print(f"    Parsed datetime: {match_datetime}")
            print(f"    Match: {match.get('game', 'Unknown')} - Date: {match_date}")
            
            # Check if match is within the desired range (today to 1 week ahead)
            if today <= match_date <= one_week_ahead:
                filtered_matches.append(match)
                print(f"      ✅ INCLUDED (within range)")
            else:
                print(f"      ❌ EXCLUDED (outside range: {match_date} not between {today} and {one_week_ahead})")
                
        except ValueError as e:
            # Skip matches with invalid date format
            print(f"    ❌ ERROR parsing date '{match_time_str}': {e}")
            print(f"    Tried formats: YYYY-MM-DD HH:MM:SS, YYYY-MM-DD, DD/MM/YYYY, MM/DD/YYYY")
            continue
    
    print(f"    Filtered from {len(matches)} to {len(filtered_matches)} matches")
    return filtered_matches

def filter_competitions_by_date(competitions):
    """
    Placeholder to maintain backward compatibility.
    Currently returns competitions unchanged.
    """
    return competitions

def parse_date_string(date_value):
    """
    Parse a date string or datetime object into a date object.
    Returns None if parsing fails.
    """
    if not date_value:
        return None

    if isinstance(date_value, datetime):
        return date_value.date()

    date_str = str(date_value).strip()
    if not date_str:
        return None

    date_formats = [
        '%d/%m/%Y',
        '%d/%m/%y',
        '%Y-%m-%d',
        '%Y/%m/%d',
        '%m/%d/%Y',
        '%m/%d/%y'
    ]

    for fmt in date_formats:
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue

    return None

def parse_time_string(time_value):
    """
    Parse a time string or datetime/time object into a time object.
    Returns None if parsing fails.
    """
    if not time_value:
        return None

    if isinstance(time_value, datetime):
        return time_value.time()

    if isinstance(time_value, time):
        return time_value

    time_str = str(time_value).strip()
    if not time_str:
        return None

    time_formats = ['%H:%M:%S', '%H:%M']
    for fmt in time_formats:
        try:
            return datetime.strptime(time_str, fmt).time()
        except ValueError:
            continue

    return None

def match_sort_key(match):
    """
    Sorting key for matches based on date, time and match ID
    """
    date_obj = parse_date_string(match.get('date_formatted'))
    time_obj = parse_time_string(match.get('time_local_formatted'))

    if date_obj is None:
        date_obj = datetime.max.date()
    if time_obj is None:
        time_obj = time.min

    match_id = str(match.get('matchId', ''))
    return (date_obj, time_obj, match_id)

def _state_path(output_path):
    """Path to the state file (next to the Excel file)."""
    return Path(output_path).parent / "football_fetch_state.json"

def load_fetch_state(output_path):
    """
    Load state that tracks which match IDs were written last run and which the user has deleted.
    Used so we update the file instead of recreating it: deleted matches are not re-added.
    """
    path = _state_path(output_path)
    if not path.exists():
        return {"last_written": {}, "deleted": {}}
    try:
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return {
            "last_written": data.get("last_written", {}),
            "deleted": data.get("deleted", {})
        }
    except Exception as e:
        print(f"⚠️  Could not load state file ({e}). Deleted-match tracking will not apply this run.")
        return {"last_written": {}, "deleted": {}}

def save_fetch_state(output_path, state):
    """Save state after writing the Excel file."""
    path = _state_path(output_path)
    try:
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(state, f, indent=2)
    except Exception as e:
        print(f"⚠️  Could not save state file ({e}).")

def load_existing_matches(output_path):
    """
    Read the previous Excel file to get any manual notes or checks that were added
    This way when we create a new file, we don't lose the work someone did manually
    We also clean up old matches that are more than a week old and already marked as complete
    """
    existing_matches = {}
    target_path = Path(output_path)

    if not target_path.exists():
        return existing_matches

    try:
        wb = openpyxl.load_workbook(output_path)
    except Exception as e:
        print(f"⚠️  Could not load existing Excel file ({e}). Continuing without previous data.")
        return existing_matches

    if "FMM automation" not in wb.sheetnames:
        return existing_matches

    ws = wb["FMM automation"]
    today = datetime.now().date()
    one_week_ago = today - timedelta(days=7)
    current_competition_id = None

    row = 1
    while row <= ws.max_row:
        cell_value = ws.cell(row=row, column=1).value

        if cell_value == 'League':
            row += 1
            continue

        if cell_value and "Competition ID" in str(cell_value):
            comp_match = re.search(r"Competition ID\s+(\d+)", str(cell_value))
            current_competition_id = comp_match.group(1) if comp_match else None
            if current_competition_id:
                existing_matches.setdefault(current_competition_id, {})
            row += 1
            continue

        if cell_value and isinstance(cell_value, str) and cell_value.startswith("Last check"):
            current_competition_id = None
            row += 1
            continue

        if current_competition_id:
            match_id_cell = ws.cell(row=row, column=7).value
            if match_id_cell:
                match_id = str(match_id_cell).strip()

                # Preserve League column (column 1) - user may have added notes there
                league_cell = ws.cell(row=row, column=1).value
                league_column_note = str(league_cell).strip() if league_cell else ''

                date_cell = ws.cell(row=row, column=2).value
                if isinstance(date_cell, datetime):
                    date_formatted = date_cell.strftime('%d/%m/%Y')
                else:
                    date_formatted = str(date_cell or '').strip()

                time_local_cell = ws.cell(row=row, column=3).value
                if isinstance(time_local_cell, datetime):
                    time_local_formatted = time_local_cell.strftime('%H:%M')
                elif isinstance(time_local_cell, time):
                    time_local_formatted = time_local_cell.strftime('%H:%M')
                else:
                    time_local_formatted = str(time_local_cell or '').strip()

                time_utc_cell = ws.cell(row=row, column=4).value
                if isinstance(time_utc_cell, datetime):
                    time_utc_formatted = time_utc_cell.strftime('%H:%M')
                elif isinstance(time_utc_cell, time):
                    time_utc_formatted = time_utc_cell.strftime('%H:%M')
                else:
                    time_utc_formatted = str(time_utc_cell or '').strip()

                time_tallinn_cell = ws.cell(row=row, column=5).value
                if isinstance(time_tallinn_cell, datetime):
                    time_tallinn_formatted = time_tallinn_cell.strftime('%H:%M')
                elif isinstance(time_tallinn_cell, time):
                    time_tallinn_formatted = time_tallinn_cell.strftime('%H:%M')
                else:
                    time_tallinn_formatted = str(time_tallinn_cell or '').strip()

                time_medellin_cell = ws.cell(row=row, column=6).value
                if isinstance(time_medellin_cell, datetime):
                    time_medellin_formatted = time_medellin_cell.strftime('%H:%M')
                elif isinstance(time_medellin_cell, time):
                    time_medellin_formatted = time_medellin_cell.strftime('%H:%M')
                else:
                    time_medellin_formatted = str(time_medellin_cell or '').strip()

                game_value = ws.cell(row=row, column=8).value
                livestream_status = str(ws.cell(row=row, column=9).value or '').strip()
                coretools_check = str(ws.cell(row=row, column=10).value or '').strip()
                whst_status = str(ws.cell(row=row, column=11).value or '').strip()
                publish_status = str(ws.cell(row=row, column=12).value or '').strip()
                webcast_status = str(ws.cell(row=row, column=13).value or '').strip()
                end_game_status = str(ws.cell(row=row, column=14).value or '').strip()

                match_date_obj = parse_date_string(date_formatted)

                if match_date_obj and match_date_obj < one_week_ago and end_game_status.lower() == 'complete':
                    print(f"    Removing completed match {match_id} (older than 1 week)")
                    row += 1
                    continue

                existing_matches[current_competition_id][match_id] = {
                    'matchId': match_id,
                    'league_column_note': league_column_note,
                    'date_formatted': date_formatted,
                    'time_local_formatted': time_local_formatted,
                    'time_utc_formatted': time_utc_formatted,
                    'time_tallinn_formatted': time_tallinn_formatted,
                    'time_medellin_formatted': time_medellin_formatted,
                    'game': str(game_value or '').strip(),
                    'livestream_status': livestream_status or 'N/A',
                    'coretools_check': coretools_check,
                    'whst_live_data_source_match': whst_status or 'N/A',
                    'publish_connection_status': publish_status or 'N/A',
                    'webcast_status': webcast_status,
                    'end_game_status': end_game_status
                }

        row += 1

    return existing_matches

def merge_matches_with_existing(new_matches, existing_matches, deleted_match_ids=None):
    """
    Combine the new match data we just fetched with the old data from the previous Excel file.
    Preserves manual notes (League column, coretools, etc.) and does not re-add matches
    that the user has deleted (if deleted_match_ids is provided).
    """
    merged_matches_map = {str(mid): existing.copy() for mid, existing in existing_matches.items()}
    # Normalize to strings so int/string mismatch (e.g. from JSON) doesn't exclude matches
    deleted_set = {str(mid).strip() for mid in (deleted_match_ids or [])}

    for match in new_matches:
        match_id = str(match.get('matchId', '')).strip()
        if not match_id:
            continue
        # Do not re-add matches the user has intentionally removed from the sheet
        if match_id in deleted_set and match_id not in existing_matches:
            continue

        existing_entry = existing_matches.get(match_id)
        merged_entry = existing_entry.copy() if existing_entry else {}

        # Update with new automated data
        merged_entry.update(match)

        # Preserve only truly manual columns; all checks (including end game) are overwritten by this run
        if existing_entry:
            merged_entry['coretools_check'] = existing_entry.get('coretools_check', '')
            merged_entry['league_column_note'] = existing_entry.get('league_column_note', '')
            # webcast_status: use new value when we have one, else keep existing
            if 'webcast_status' not in match or not match.get('webcast_status'):
                merged_entry['webcast_status'] = existing_entry.get('webcast_status', '')
        else:
            merged_entry.setdefault('coretools_check', '')
            merged_entry.setdefault('league_column_note', '')
            merged_entry.setdefault('end_game_status', match.get('end_game_status', ''))
            merged_entry.setdefault('webcast_status', match.get('webcast_status', ''))

        merged_entry.setdefault('livestream_status', match.get('livestream_status', 'N/A'))
        merged_entry.setdefault('whst_live_data_source_match', match.get('whst_live_data_source_match', 'N/A'))
        merged_entry.setdefault('publish_connection_status', match.get('publish_connection_status', 'N/A'))
        merged_entry.setdefault('time_utc_formatted', match.get('time_utc_formatted', ''))
        merged_entry.setdefault('time_tallinn_formatted', match.get('time_tallinn_formatted', ''))
        merged_entry.setdefault('time_medellin_formatted', match.get('time_medellin_formatted', ''))
        merged_entry.setdefault('game', match.get('game', ''))
        merged_entry.setdefault('date_formatted', match.get('date_formatted', ''))
        merged_entry.setdefault('time_local_formatted', match.get('time_local_formatted', ''))

        merged_matches_map[match_id] = merged_entry

    merged_list = list(merged_matches_map.values())
    merged_list.sort(key=match_sort_key)
    return merged_list

def get_federation_code_from_match(api_key, match_id, match_data=None):
    """
    Try to get federation code (league abbreviation) from match data.
    If match_data is provided, uses it instead of making an API call.
    Returns federation code or None if not found.
    """
    if not api_key or not match_id:
        return None
    
    # Use provided match_data or fetch it (will use cache)
    if match_data is None:
        _, _, match_data = fetch_match_details(api_key, match_id, return_full_data=True)
    
    if not match_data:
        return None
    
    # Try to get league abbreviation from match data (same keys as fetch_league_details)
    league_obj = match_data.get('league') or {}
    league_abbrev = (
        match_data.get('leagueAbbrev') or match_data.get('abbreviation') or
        match_data.get('federationCode') or match_data.get('code') or match_data.get('slug') or
        league_obj.get('leagueAbbrev') or league_obj.get('abbreviation') or
        league_obj.get('federationCode') or league_obj.get('code') or league_obj.get('slug')
    )
    
    # If not found, try to get league ID and fetch it
    if not league_abbrev:
        league_id = match_data.get('leagueId') or (match_data.get('league') or {}).get('leagueId')
        if league_id:
            league_abbrev = fetch_league_details(api_key, league_id)
    
    return league_abbrev

def fetch_match_details(api_key, match_id, return_full_data=False):
    """
    Get information about a specific match - whether it has livestream and what data source it uses
    Uses caching to avoid fetching the same match multiple times
    Can optionally return the full match data if we need it for other checks
    """
    # Check if we already fetched this match's details before
    cache_key = f"{match_id}"
    if cache_key in _match_details_cache:
        cached = _match_details_cache[cache_key]
        if return_full_data:
            return cached[0], cached[1], cached[2]
        return cached[0], cached[1]
    
    url = f"https://api.wh.geniussports.com/v1/football/matches/{match_id}"
    params = {
        'ak': api_key
    }
    
    try:
        response = requests.get(url, params=params, timeout=30)
        
        if response.status_code != 200:
            print(f"        ⚠️  Could not fetch match details for match {match_id} (HTTP {response.status_code})")
            result = ("N/A", None, None) if return_full_data else ("N/A", None)
            _match_details_cache[cache_key] = ("N/A", None, None)
            return result
        
        data = response.json()
        
        # Extract match data from response
        match_data = None
        if 'response' in data:
            if 'data' in data['response']:
                match_data = data['response']['data']
            else:
                match_data = data['response']
        else:
            match_data = data
        
        if not match_data:
            print(f"        ⚠️  Could not extract match data for match {match_id}")
            result = ("N/A", None, None) if return_full_data else ("N/A", None)
            _match_details_cache[cache_key] = ("N/A", None, None)
            return result
        
        # Extract liveStream field
        live_stream_value = match_data.get('liveStream')
        
        # Convert to Yes/No
        if live_stream_value == 1:
            livestream_status = "Yes"
        elif live_stream_value == 0:
            livestream_status = "No"
        else:
            print(f"        ⚠️  liveStream field not found or has unexpected value for match {match_id}")
            livestream_status = "N/A"
        
        # Extract LiveDataSource field
        # Try different possible locations for LiveDataSource
        live_data_source = (
            match_data.get('LiveDataSource') or
            (match_data.get('internalConfiguration') or {}).get('LiveDataSource') or
            match_data.get('competitionLiveDataSource') or
            match_data.get('liveDataSource')
        )
        
        # Cache the result
        _match_details_cache[cache_key] = (livestream_status, live_data_source, match_data)
        
        if return_full_data:
            return (livestream_status, live_data_source, match_data)
        return (livestream_status, live_data_source)
            
    except requests.exceptions.RequestException as e:
        print(f"        ⚠️  API call failed for match {match_id}: {str(e)}")
        result = ("N/A", None, None) if return_full_data else ("N/A", None)
        _match_details_cache[cache_key] = ("N/A", None, None)
        return result
    except Exception as e:
        print(f"        ⚠️  Error processing match details for {match_id}: {str(e)}")
        result = ("N/A", None, None) if return_full_data else ("N/A", None)
        _match_details_cache[cache_key] = ("N/A", None, None)
        return result

def check_publish_connection(api_key, match_id):
    """
    Check if there's a publish connection set up for this match
    Publish connections are needed to send match data to external systems
    Returns "Yes" if there's at least one connection, "No" if none, "N/A" if we can't check
    """
    url = "https://api.wh.geniussports.com/v1/football/connections"
    params = {
        'ak': api_key,
        'matchId': match_id,
        'type': 'publish'
    }
    
    try:
        response = requests.get(url, params=params, timeout=30)
        
        if response.status_code != 200:
            print(f"        ⚠️  Could not fetch publish connections for match {match_id} (HTTP {response.status_code})")
            return "N/A"
        
        data = response.json()
        
        # Extract data array from response
        connections_data = []
        if 'response' in data:
            if 'data' in data['response']:
                connections_data = data['response']['data']
            else:
                connections_data = []
        else:
            connections_data = []
        
        # Check if there's at least one connection
        if isinstance(connections_data, list) and len(connections_data) >= 1:
            return "Yes"
        else:
            return "No"
            
    except requests.exceptions.RequestException as e:
        print(f"        ⚠️  API call failed for publish connections (match {match_id}): {str(e)}")
        return "N/A"
    except Exception as e:
        print(f"        ⚠️  Error processing publish connections for match {match_id}: {str(e)}")
        return "N/A"

def check_webcast_data(match_id):
    """
    Check if the webcast page has all the data it needs before a match starts
    We need team names and player lineups for both teams
    This is important because if the webcast doesn't have this data, viewers won't see proper information
    """
    if not match_id:
        return "N/A"
    
    try:
        url = f"https://livestats.dcd.shared.geniussports.com/data/football/{match_id}/data.json"
        response = requests.get(url, timeout=30)
        
        if response.status_code == 403:
            print(f"        ℹ️  Webcast JSON not yet available for match {match_id} (HTTP 403 - typically becomes available closer to match time)")
            return "N/A"
        elif response.status_code == 404:
            print(f"        ℹ️  Webcast JSON not found for match {match_id} (HTTP 404 - may not be available yet)")
            return "N/A"
        elif response.status_code != 200:
            print(f"        ⚠️  Could not fetch webcast data for match {match_id} (HTTP {response.status_code})")
            return "N/A"
        
        data = response.json()
        
        # Check for team names in 'tm' object (team metadata)
        # Team names can be stored in various locations and field names
        tm = data.get('tm', {})
        team1_name = None
        team2_name = None
        
        # Possible field names for team names
        name_fields = ['name', 'nameInternational', 'teamName', 'teamNameInternational', 'n', 'nm', 'team', 'teamNameInt']
        
        if isinstance(tm, dict):
            # Try all possible keys for team 1 (string, int, and other variations)
            for key in list(tm.keys()):
                if str(key) in ['1', 'team1', 'home', '0'] or key == 1:
                    if isinstance(tm[key], dict):
                        # Try all possible name field variations
                        for field in name_fields:
                            team1_name = tm[key].get(field)
                            if team1_name:
                                break
                        if team1_name:
                            break
            
            # Try all possible keys for team 2 (string, int, and other variations)
            for key in list(tm.keys()):
                if str(key) in ['2', 'team2', 'away', '1'] or key == 2:
                    if isinstance(tm[key], dict):
                        # Try all possible name field variations
                        for field in name_fields:
                            team2_name = tm[key].get(field)
                            if team2_name:
                                break
                        if team2_name:
                            break
        
        # If still not found, try checking all keys in tm
        if not team1_name or not team2_name:
            if isinstance(tm, dict):
                # Get all keys and try to find team data
                all_keys = list(tm.keys())
                for key in all_keys:
                    if isinstance(tm[key], dict):
                        team_obj = tm[key]
                        # Check if this looks like team data (has name fields or player data)
                        has_name = any(field in team_obj for field in name_fields)
                        has_players = 'pl' in team_obj or 'players' in team_obj
                        
                        if has_name or has_players:
                            # Try to determine which team this is
                            # If we don't have team1 yet, assign it
                            if not team1_name:
                                for field in name_fields:
                                    team1_name = team_obj.get(field)
                                    if team1_name:
                                        break
                            # If we have team1 but not team2, and this is different, assign it
                            elif not team2_name and team1_name != team_obj.get('name'):
                                for field in name_fields:
                                    team2_name = team_obj.get(field)
                                    if team2_name:
                                        break
                            
                            if team1_name and team2_name:
                                break
        
        has_team1_name = bool(team1_name)
        has_team2_name = bool(team2_name)
        
        # Check for players data in 'tm' object (primary location)
        # Players are stored in tm['1']['pl'] and tm['2']['pl']
        # Each player has 'starter' (1 for starters) and 'active' fields
        team1_players = []
        team2_players = []
        
        if isinstance(tm, dict):
            # Collect all team objects from tm
            all_team_objects = []
            for key in list(tm.keys()):
                if isinstance(tm[key], dict):
                    team_obj = tm[key]
                    # Check if this looks like team data (has players or name fields)
                    has_players = any(field in team_obj for field in ['pl', 'players', 'player', 'p'])
                    has_name = any(field in team_obj for field in name_fields)
                    if has_players or has_name:
                        all_team_objects.append(team_obj)
            
            # Extract players from team objects
            for team_obj in all_team_objects:
                # Try different player field names
                for pl_field in ['pl', 'players', 'player', 'p']:
                    team_pl = team_obj.get(pl_field)
                    if isinstance(team_pl, dict):
                        players = list(team_pl.values())
                        # Assign to team1 or team2 based on which one is empty
                        if not team1_players:
                            team1_players = players
                        elif not team2_players and players != team1_players:
                            team2_players = players
                        break
                    elif isinstance(team_pl, list):
                        # Assign to team1 or team2 based on which one is empty
                        if not team1_players:
                            team1_players = team_pl
                        elif not team2_players and team_pl != team1_players:
                            team2_players = team_pl
                        break
                if team1_players and team2_players:
                    break
        
        # Also check 'teams' object as fallback (some matches might use this structure)
        # The teams object contains player data keyed by player IDs with 'tno' field
        teams = data.get('teams')
        if (not team1_players or not team2_players) and isinstance(teams, dict) and len(teams) > 0:
            # Only use teams object if we didn't find players in tm['pl']
            # Handle both string and integer tno values
            if not team1_players:
                team1_players = [p for p in teams.values() if isinstance(p, dict) and (p.get('tno') == 1 or p.get('tno') == '1')]
            if not team2_players:
                team2_players = [p for p in teams.values() if isinstance(p, dict) and (p.get('tno') == 2 or p.get('tno') == '2')]
        
        # Filter to only dict objects
        team1_players = [p for p in team1_players if isinstance(p, dict)]
        team2_players = [p for p in team2_players if isinstance(p, dict)]
        
        # Check if we have players for both teams
        has_team1_players = len(team1_players) > 0
        has_team2_players = len(team2_players) > 0
        
        # Check if we have lineups (starters) for both teams
        team1_starters = [p for p in team1_players if p.get('starter') == 1]
        team2_starters = [p for p in team2_players if p.get('starter') == 1]
        
        has_team1_lineup = len(team1_starters) > 0
        has_team2_lineup = len(team2_starters) > 0
        
        # Evaluate the check:
        # - We need both team names (from 'tm' object)
        # - We need both teams to have players in the 'teams' object
        # - Ideally, both teams should have starters (lineups)
        
        if not has_team1_name or not has_team2_name:
            print(f"        Match {match_id}: Webcast missing team names (Team 1: {has_team1_name}, Team 2: {has_team2_name})")
            return "No"
        
        if not has_team1_players or not has_team2_players:
            print(f"        Match {match_id}: Webcast missing players for one or both teams (Team 1: {len(team1_players)}, Team 2: {len(team2_players)})")
            return "No"
        
        # If we have both team names and players, check for lineups
        if has_team1_lineup and has_team2_lineup:
            print(f"        Match {match_id}: Webcast has both teams with names and lineups (Team 1: {team1_name}, Team 2: {team2_name})")
            return "Yes"
        else:
            # Has team names and players, but lineups may not be set yet (match hasn't started)
            print(f"        Match {match_id}: Webcast has team names and players but lineups not set (Team 1: {team1_name}, Team 2: {team2_name}, Starters: T1={len(team1_starters)}, T2={len(team2_starters)})")
            # Still return "Yes" if we have team names and players, as lineups may be set closer to match time
            return "Yes"
            
    except requests.exceptions.RequestException as e:
        print(f"        ⚠️  API call failed for webcast data (match {match_id}): {str(e)}")
        return "N/A"
    except Exception as e:
        print(f"        ⚠️  Error processing webcast data for match {match_id}: {str(e)}")
        return "N/A"

def fetch_unregistered_players(api_key, match_id):
    """
    Get a list of players who played in the match but weren't registered in the system
    These players might have actions (goals, cards) assigned to them, which is why we need to check
    """
    if not api_key or not match_id:
        return []
    
    try:
        url = f"https://api.wh.geniussports.com/v1/football/matches/{match_id}/persons/unregistered"
        params = {'ak': api_key}
        response = requests.get(url, params=params, timeout=30)
        
        if response.status_code != 200:
            return []
        
        data = response.json()
        unregistered_players = []
        
        if 'response' in data and 'data' in data['response']:
            unregistered_players = data['response']['data']
        elif 'data' in data:
            unregistered_players = data['data']
        
        return unregistered_players if isinstance(unregistered_players, list) else []
        
    except Exception as e:
        print(f"        ⚠️  Error fetching unregistered players for match {match_id}: {e}")
        return []

def check_end_game_past_match_data(match_id, federation_code, api_key=None):
    """
    After a match is finished, check if all the data is complete and correct.
    Returns:
      "Yes" - all good (teams, scores, time, and all events assigned to players; subs balanced).
      "No"  - issues found (unassigned events, incorrect lineup, or unregistered player problems).
              Caller turns this into "Match check required" + HS link in Excel/Slack.
      "N/A" - check could not run: no federation_code; HS page 404/403/other HTTP error; or exception.
    """
    if not match_id or not federation_code:
        return "N/A"
    
    try:
        # HS URLs use lowercase federation path (e.g. unafut not UNAFUT)
        fed_slug = str(federation_code).strip().lower()
        url = f"https://hosted.dcd.shared.geniussports.com/{fed_slug}/en/match/{match_id}/summary?json=1"
        response = requests.get(url, timeout=30)
        
        if response.status_code == 404:
            print(f"        ℹ️  HS JSON not found for match {match_id} (HTTP 404). Tried: {url}")
            return "N/A"
        elif response.status_code == 403:
            print(f"        ℹ️  HS JSON not accessible for match {match_id} (HTTP 403)")
            return "N/A"
        elif response.status_code != 200:
            print(f"        ⚠️  Could not fetch HS JSON for match {match_id} (HTTP {response.status_code}). Tried: {url}")
            return "N/A"
        
        data = response.json()
        
        # Check for team names from matchDetail.competitors
        team1_name = None
        team2_name = None
        team1_id = None
        team2_id = None
        
        # Try matchDetail.competitors first (primary location based on JSON structure)
        match_detail = data.get('matchDetail', {})
        if isinstance(match_detail, dict):
            competitors = match_detail.get('competitors', [])
            if isinstance(competitors, list):
                for competitor in competitors:
                    if isinstance(competitor, dict):
                        is_home = competitor.get('isHomeCompetitor', 0)
                        comp_name = competitor.get('competitorName') or competitor.get('teamName')
                        team_id = competitor.get('teamId') or competitor.get('competitorId')
                        
                        if comp_name:
                            if is_home == 1:
                                team1_name = comp_name
                                team1_id = team_id
                            else:
                                team2_name = comp_name
                                team2_id = team_id
        
        # Fallback: try top-level competitors array
        if not team1_name or not team2_name:
            if 'competitors' in data and isinstance(data['competitors'], list):
                for competitor in data['competitors']:
                    if isinstance(competitor, dict):
                        is_home = competitor.get('isHomeCompetitor', 0)
                        comp_name = (
                            competitor.get('competitorName') or
                            competitor.get('teamName') or
                            competitor.get('teamNameInternational') or
                            competitor.get('name')
                        )
                        team_id = competitor.get('teamId') or competitor.get('competitorId')
                        if comp_name:
                            if is_home == 1:
                                team1_name = team1_name or comp_name
                                team1_id = team1_id or team_id
                            else:
                                team2_name = team2_name or comp_name
                                team2_id = team2_id or team_id
        
        has_team_names = bool(team1_name and team2_name)
        
        # Check for time/date from matchDetail.matchTime
        has_time = False
        if isinstance(match_detail, dict):
            match_time = match_detail.get('matchTime') or match_detail.get('matchTimeUTC')
            if match_time:
                has_time = True
        
        # Fallback: check top-level
        if not has_time:
            has_time = any(key in data for key in ['matchTime', 'time', 'date', 'startTime', 'matchDate', 'kickoffTime'])
        
        # Check for scores from matchDetail.competitors[].scoreString
        has_scores = False
        scores_found = []
        
        if isinstance(match_detail, dict):
            competitors = match_detail.get('competitors', [])
            if isinstance(competitors, list):
                for competitor in competitors:
                    if isinstance(competitor, dict):
                        score_str = competitor.get('scoreString')
                        score_val = competitor.get('score')
                        if score_str is not None and score_str != '':
                            scores_found.append(score_str)
                        elif score_val is not None:
                            scores_found.append(score_val)
                has_scores = len(scores_found) >= 2
        
        # Fallback: try top-level competitors
        if not has_scores:
            if 'competitors' in data and isinstance(data['competitors'], list):
                for competitor in data['competitors']:
                    if isinstance(competitor, dict):
                        score_str = competitor.get('scoreString')
                        score_val = competitor.get('score')
                        if score_str is not None and score_str != '':
                            scores_found.append(score_str)
                        elif score_val is not None:
                            scores_found.append(score_val)
                has_scores = len(scores_found) >= 2
        
        # Try other locations if still not found
        if not has_scores:
            if 'score' in data:
                score = data['score']
                if isinstance(score, dict):
                    has_scores = bool(score.get('home') is not None and score.get('away') is not None)
                elif isinstance(score, str):
                    has_scores = bool(score)
            elif 'scores' in data:
                scores = data['scores']
                if isinstance(scores, dict):
                    has_scores = bool(scores.get('home') is not None and scores.get('away') is not None)
            elif 'homeScore' in data and 'awayScore' in data:
                has_scores = bool(data['homeScore'] is not None and data['awayScore'] is not None)
        
        # Check actions by comparing summary stats with individual player stats
        actions_valid = True
        actions_issues = []
        
        # Fetch unregistered players if API key is provided
        unregistered_players = []
        unregistered_by_team = {}  # Maps team_id to list of unregistered players
        if api_key:
            unregistered_players = fetch_unregistered_players(api_key, match_id)
            if unregistered_players:
                for unreg_player in unregistered_players:
                    if isinstance(unreg_player, dict):
                        unreg_team_id = unreg_player.get('teamId')
                        if unreg_team_id:
                            if unreg_team_id not in unregistered_by_team:
                                unregistered_by_team[unreg_team_id] = []
                            unregistered_by_team[unreg_team_id].append(unreg_player)
                print(f"        Match {match_id}: Found {len(unregistered_players)} unregistered players")
        
        # Get summary stats from comparisonStats array
        comparison_stats = data.get('comparisonStats', [])
        if not isinstance(comparison_stats, list):
            comparison_stats = []
        
        # Get individual player stats from boxscore arrays
        boxscore_home = data.get('boxscore_hometeam', [])
        boxscore_away = data.get('boxscore_awayteam', [])
        if not isinstance(boxscore_home, list):
            boxscore_home = []
        if not isinstance(boxscore_away, list):
            boxscore_away = []
        
        # Build a mapping of team IDs to home/away status from competitors
        team_to_boxscore = {}  # Maps team_id to boxscore array
        
        # Try to match teams by ID from matchDetail.competitors
        if isinstance(match_detail, dict):
            competitors = match_detail.get('competitors', [])
            if isinstance(competitors, list):
                for idx, competitor in enumerate(competitors):
                    if isinstance(competitor, dict):
                        comp_team_id = competitor.get('teamId') or competitor.get('competitorId')
                        is_home = competitor.get('isHomeCompetitor', 0)
                        if comp_team_id:
                            # Use home boxscore if isHomeCompetitor == 1, otherwise away
                            team_to_boxscore[comp_team_id] = boxscore_home if is_home == 1 else boxscore_away
        
        # Fallback: if we have team IDs from earlier, use those
        if team1_id and team1_id not in team_to_boxscore:
            team_to_boxscore[team1_id] = boxscore_home
        if team2_id and team2_id not in team_to_boxscore:
            team_to_boxscore[team2_id] = boxscore_away
        
        # Process each team's summary stats and compare with individual player stats
        for team_summary in comparison_stats:
            if not isinstance(team_summary, dict):
                continue
            
            team_id = team_summary.get('teamId')
            if not team_id:
                continue
            
            # Get summary totals
            summary_goals = team_summary.get('sGoals', 0)
            summary_yellow_cards = team_summary.get('sYellowCards', 0)
            summary_red_cards = team_summary.get('sRedCards', 0)
            
            # Determine which boxscore array to use
            boxscore_players = team_to_boxscore.get(team_id)
            
            # If we couldn't match by ID, try to find by teamId in boxscore arrays
            if not boxscore_players:
                # Check if this team_id appears in home boxscore
                found_in_home = any(p.get('teamId') == team_id for p in boxscore_home if isinstance(p, dict))
                if found_in_home:
                    boxscore_players = boxscore_home
                else:
                    # Check away boxscore
                    found_in_away = any(p.get('teamId') == team_id for p in boxscore_away if isinstance(p, dict))
                    if found_in_away:
                        boxscore_players = boxscore_away
            
            # If still not found, skip this team (can't validate)
            if not boxscore_players:
                continue
            
            # Sum up individual player stats
            player_goals = 0
            player_yellow_cards = 0
            player_red_cards = 0
            
            for player in boxscore_players:
                if isinstance(player, dict):
                    # Only count stats for players from this team
                    player_team_id = player.get('teamId')
                    if player_team_id == team_id:
                        player_goals += player.get('sGoals', 0)
                        player_yellow_cards += player.get('sYellowCards', 0)
                        player_red_cards += player.get('sRedCards', 0)
            
            # Get number of unregistered players for this team
            unreg_count = len(unregistered_by_team.get(team_id, []))
            
            # Compare: if summary has more than individual players, some actions are missing players
            # Account for unregistered players - they might have actions assigned but won't show in boxscore
            goals_discrepancy = summary_goals - player_goals
            yellow_discrepancy = summary_yellow_cards - player_yellow_cards
            red_discrepancy = summary_red_cards - player_red_cards
            
            if goals_discrepancy > 0:
                if unreg_count > 0:
                    # Unregistered players might account for some discrepancy, but we can't verify
                    actions_issues.append(f"Team {team_id}: {summary_goals} goals in summary but only {player_goals} assigned to registered players ({unreg_count} unregistered players may account for discrepancy)")
                    # Still flag as invalid since we can't verify unregistered players have the actions
                    actions_valid = False
                else:
                    actions_valid = False
                    actions_issues.append(f"Team {team_id}: {summary_goals} goals in summary but only {player_goals} assigned to players")
            
            if yellow_discrepancy > 0:
                if unreg_count > 0:
                    actions_issues.append(f"Team {team_id}: {summary_yellow_cards} yellow cards in summary but only {player_yellow_cards} assigned to registered players ({unreg_count} unregistered players may account for discrepancy)")
                    actions_valid = False
                else:
                    actions_valid = False
                    actions_issues.append(f"Team {team_id}: {summary_yellow_cards} yellow cards in summary but only {player_yellow_cards} assigned to players")
            
            if red_discrepancy > 0:
                if unreg_count > 0:
                    actions_issues.append(f"Team {team_id}: {summary_red_cards} red cards in summary but only {player_red_cards} assigned to registered players ({unreg_count} unregistered players may account for discrepancy)")
                    actions_valid = False
                else:
                    actions_valid = False
                    actions_issues.append(f"Team {team_id}: {summary_red_cards} red cards in summary but only {player_red_cards} assigned to players")
        
        # Check substitutions: every player subbed off should have a corresponding player subbed on at the same time
        # Process all teams from both boxscore arrays
        all_teams_processed = set()
        for boxscore_array in [boxscore_home, boxscore_away]:
            for player in boxscore_array:
                if not isinstance(player, dict):
                    continue
                player_team_id = player.get('teamId')
                if not player_team_id or player_team_id in all_teams_processed:
                    continue
                all_teams_processed.add(player_team_id)
                
                # Collect players subbed off and subbed on for this team
                players_subbed_off = []  # List of (time, player_id) tuples
                players_subbed_on = []    # List of (time, player_id) tuples
                
                for p in boxscore_array:
                    if not isinstance(p, dict):
                        continue
                    if p.get('teamId') != player_team_id:
                        continue
                    
                    sub_off_time = p.get('sSubstitutionOffTime', '0')
                    sub_on_time = p.get('sSubstitutionOnTime', '0')
                    
                    # Convert to string and check if it's a valid time (not '0' or empty)
                    sub_off_time_str = str(sub_off_time).strip()
                    sub_on_time_str = str(sub_on_time).strip()
                    
                    # Only consider non-zero times
                    if sub_off_time_str and sub_off_time_str != '0' and sub_off_time_str != '':
                        players_subbed_off.append((sub_off_time_str, p.get('personId')))
                    
                    if sub_on_time_str and sub_on_time_str != '0' and sub_on_time_str != '':
                        players_subbed_on.append((sub_on_time_str, p.get('personId')))
                
                # Group substitutions by time
                subs_off_by_time = defaultdict(list)
                subs_on_by_time = defaultdict(list)
                
                for time_str, player_id in players_subbed_off:
                    subs_off_by_time[time_str].append(player_id)
                
                for time_str, player_id in players_subbed_on:
                    subs_on_by_time[time_str].append(player_id)
                
                # Check that for each substitution time, the number of players subbed off equals the number subbed on
                all_sub_times = set(list(subs_off_by_time.keys()) + list(subs_on_by_time.keys()))
                
                for sub_time in all_sub_times:
                    count_off = len(subs_off_by_time[sub_time])
                    count_on = len(subs_on_by_time[sub_time])
                    
                    if count_off != count_on:
                        actions_valid = False
                        if count_off > count_on:
                            actions_issues.append(f"Team {player_team_id}: At minute {sub_time}, {count_off} player(s) subbed off but only {count_on} player(s) subbed on")
                        else:
                            actions_issues.append(f"Team {player_team_id}: At minute {sub_time}, {count_on} player(s) subbed on but only {count_off} player(s) subbed off")
        
        # If no comparisonStats found, try alternative validation
        if not comparison_stats and (boxscore_home or boxscore_away):
            # If we have boxscore data but no comparison stats, we can't validate
            # This might be okay if the match structure is different
            actions_valid = True
        elif not comparison_stats and not boxscore_home and not boxscore_away:
            # No stats data at all - might be a match that hasn't been fully processed
            actions_valid = True
        
        # Evaluate: all checks must pass
        if not has_team_names:
            print(f"        Match {match_id}: HS JSON missing team names (Team 1: {team1_name}, Team 2: {team2_name})")
            return "No"
        
        if not has_time:
            print(f"        Match {match_id}: HS JSON missing time/date")
            return "No"
        
        if not has_scores:
            print(f"        Match {match_id}: HS JSON missing scores")
            return "No"
        
        if not actions_valid:
            if actions_issues:
                print(f"        Match {match_id}: HS JSON has actions without players assigned:")
                for issue in actions_issues[:3]:  # Show first 3 issues
                    print(f"        Match {match_id}:   - {issue}")
            else:
                print(f"        Match {match_id}: HS JSON has actions without players assigned (validation failed)")
            return "No"
        
        print(f"        Match {match_id}: HS JSON has all required past match data (teams: {team1_name} vs {team2_name})")
        return "Yes"
            
    except requests.exceptions.RequestException as e:
        print(f"        ⚠️  API call failed for HS JSON (match {match_id}): {str(e)}")
        return "N/A"
    except Exception as e:
        print(f"        ⚠️  Error processing HS JSON for match {match_id}: {str(e)}")
        return "N/A"

def fetch_matches_for_competition(api_key, competition_id, competition_name, competition_live_data_source=None, league_abbrev=None):
    """
    Fetch matches for a specific competition (1 week ago to 1 week ahead)
    """
    print(f"    Fetching matches for competition: {competition_name} (ID: {competition_id})")
    
    # Get date range for API call - include past matches (1 week ago) to upcoming (1 week ahead)
    today = datetime.now().date()
    one_week_ahead = today + timedelta(days=7)
    one_week_ago = today - timedelta(days=7)
    
    # Fetch matches from 1 week ago to 1 week ahead
    url = f"https://api.wh.geniussports.com/v1/football/competitions/{competition_id}/matches"
    params = {
        'ak': api_key,
        'fromDate': one_week_ago.strftime('%Y-%m-%d'),
        'toDate': one_week_ahead.strftime('%Y-%m-%d'),
        'limit': 500
    }
    
    print(f"    API call with date range: {one_week_ago} to {one_week_ahead} (includes past matches for end game check)")
    
    try:
        response = requests.get(url, params=params, timeout=30)
        
        if response.status_code != 200:
            print(f"      API returned error status: {response.status_code}")
            return []
        
        data = response.json()
        
        # Extract matches from response
        all_matches = []
        if 'response' in data and 'data' in data['response']:
            matches_data = data['response']['data']
            print(f"      Found {len(matches_data)} total matches")
            
            for match in matches_data:
                match_info = {
                    'matchId': match.get('matchId', ''),
                    'matchTime': match.get('matchTime', ''),  # Format: YYYY-MM-DD HH:MM:SS local time
                    'matchTimeUTC': match.get('matchTimeUTC', ''),  # Format: YYYY-MM-DD HH:MM:SS UTC time
                    'competitor1': '',
                    'competitor2': '',
                    'competition': competition_name,
                    'venue': match.get('venue', {}).get('venueName', '') if match.get('venue') else '',
                    'status': match.get('matchStatus', '')
                }
                
                # Extract both competitor names from the competitors array
                if 'competitors' in match and isinstance(match['competitors'], list):
                    competitors = match['competitors']
                    if len(competitors) >= 2:
                        match_info['competitor1'] = competitors[0].get('teamNameInternational', '') or competitors[0].get('competitorName', '')
                        match_info['competitor2'] = competitors[1].get('teamNameInternational', '') or competitors[1].get('competitorName', '')
                
                all_matches.append(match_info)
        
        # No need to filter by date since API already returns matches in the date range
        # Format the matches for Excel
        matches = []
        for match_info in all_matches:
            formatted_match = format_match_data_for_excel(match_info)
            matches.append(formatted_match)
            print(f"        - {formatted_match['game']} (ID: {formatted_match['matchId']}) - {formatted_match['date_formatted']} {formatted_match['time_local_formatted']}")
        
        print(f"      Found {len(matches)} matches (API already filtered by date range)")
        
        # Process matches in parallel for better performance
        print(f"      Processing {len(matches)} matches in parallel...")
        
        def process_single_match(match, api_key, competition_live_data_source, league_abbrev):
            """
            Process all the checks for a single match
            This function can run in parallel with other matches to speed things up
            It does all the validation checks: livestream, data source, connections, webcast, and end game data
            """
            match_id = match.get('matchId', '')
            if not match_id:
                # If no match ID, set everything to N/A
                match['livestream_status'] = "N/A"
                match['whst_live_data_source_match'] = "N/A"
                match['publish_connection_status'] = "N/A"
                match['webcast_status'] = "N/A"
                match['end_game_status'] = ''
                return match
            
            # Get match details - we ask for full data so we can reuse it later (saves an API call)
            livestream_status, match_live_data_source, match_data = fetch_match_details(api_key, match_id, return_full_data=True)
            match['livestream_status'] = livestream_status
            match['match_live_data_source'] = match_live_data_source
            
            # Check if the match's data source matches what the competition expects
            # This is important for data quality - they should match
            if competition_live_data_source and match_live_data_source:
                comp_ds = str(competition_live_data_source).strip()
                match_ds = str(match_live_data_source).strip()
                if comp_ds.lower() == match_ds.lower():
                    match['whst_live_data_source_match'] = "Yes"
                else:
                    match['whst_live_data_source_match'] = "No"
            else:
                match['whst_live_data_source_match'] = "N/A"
            
            # Check if there are publish connections set up for this match
            publish_connection_status = check_publish_connection(api_key, match_id)
            match['publish_connection_status'] = publish_connection_status
            
            # Check if the webcast has all the data it needs (team names, lineups)
            webcast_status = check_webcast_data(match_id)
            match['webcast_status'] = webcast_status
            
            # Run HS end-game check only when kickoff was at least 2 hours ago (match has had time to finish).
            try:
                utc_str = (match.get('matchTimeUTC') or '').strip()
                local_str = (match.get('matchTime') or '').strip()
                kickoff_plus_2h = None
                use_utc = False
                if utc_str:
                    try:
                        kickoff_utc = datetime.strptime(utc_str, '%Y-%m-%d %H:%M:%S').replace(tzinfo=timezone.utc)
                        kickoff_plus_2h = kickoff_utc + timedelta(hours=2)
                        use_utc = True
                    except ValueError:
                        pass
                if kickoff_plus_2h is None and local_str:
                    try:
                        kickoff_local = datetime.strptime(local_str, '%Y-%m-%d %H:%M:%S')
                        kickoff_plus_2h = kickoff_local + timedelta(hours=2)
                    except ValueError:
                        pass
                now = datetime.now(timezone.utc) if use_utc else datetime.now()
                if kickoff_plus_2h is not None and kickoff_plus_2h <= now:
                    federation_code = league_abbrev
                    if not federation_code:
                        federation_code = get_federation_code_from_match(api_key, match_id, match_data)
                    if federation_code:
                        print(f"        Match {match_id}: Running HS end-game check...")
                        end_game_status = check_end_game_past_match_data(match_id, federation_code, api_key)
                        if end_game_status == "No":
                                match['end_game_status'] = "Match check required"
                                match['end_game_hs_url'] = f"https://hosted.dcd.shared.geniussports.com/{str(federation_code).strip().lower()}/en/match/{match_id}/summary"
                        elif end_game_status == "N/A":
                            match['end_game_status'] = "N/A - Match check required"
                            match['end_game_hs_url'] = f"https://hosted.dcd.shared.geniussports.com/{str(federation_code).strip().lower()}/en/match/{match_id}/summary"
                        else:
                            match['end_game_status'] = end_game_status  # Yes
                    else:
                        print(f"        Match {match_id}: HS check N/A (no federation code for league)")
                        match['end_game_status'] = "N/A - Match check required"
                else:
                    if kickoff_plus_2h is None:
                        print(f"        Match {match_id}: Skipping HS end-game check (no kickoff time)")
                    else:
                        print(f"        Match {match_id}: Skipping HS end-game check (kickoff < 2h ago)")
                    match['end_game_status'] = 'Too early'
            except Exception as e:
                match['end_game_status'] = 'Too early'
            
            return match
        
        # Process all matches at the same time (in parallel) instead of one by one
        # This is much faster! We limit to 10 at a time so we don't overwhelm the API
        max_workers = min(10, len(matches))  # Don't do more than 10 matches at once
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            future_to_match = {
                executor.submit(process_single_match, match, api_key, competition_live_data_source, league_abbrev): match
                for match in matches
            }
            
            processed_matches = []
            for future in as_completed(future_to_match):
                try:
                    processed_match = future.result()
                    processed_matches.append(processed_match)
                    match_id = processed_match.get('matchId', '')
                    print(f"        Match {match_id}: Completed processing")
                except Exception as e:
                    match = future_to_match[future]
                    match_id = match.get('matchId', '')
                    print(f"        ⚠️  Error processing match {match_id}: {e}")
                    # Set defaults on error
                    match['livestream_status'] = "N/A"
                    match['whst_live_data_source_match'] = "N/A"
                    match['publish_connection_status'] = "N/A"
                    match['webcast_status'] = "N/A"
                    match['end_game_status'] = ''
                    processed_matches.append(match)
        
        matches = processed_matches
        
        # Include all matches from 1 week ago to 1 week ahead
        # This allows us to see end game status for completed matches and upcoming matches
        today = datetime.now().date()
        one_week_ahead = today + timedelta(days=7)
        one_week_ago = today - timedelta(days=7)
        filtered_matches = []
        
        for match in matches:
            match_time_str = match.get('matchTime', '')
            if match_time_str:
                try:
                    match_datetime = datetime.strptime(match_time_str, '%Y-%m-%d %H:%M:%S')
                    match_date = match_datetime.date()
                    # Include matches from 1 week ago to 1 week ahead (to show past matches with end game status)
                    if one_week_ago <= match_date <= one_week_ahead:
                        filtered_matches.append(match)
                except:
                    # If date parsing fails, include the match anyway
                    filtered_matches.append(match)
            else:
                # If no match time, include it
                filtered_matches.append(match)
        
        print(f"      Filtered to {len(filtered_matches)} matches (from {one_week_ago} to {one_week_ahead}, includes past matches for end game check)")
        return filtered_matches
        
    except requests.exceptions.RequestException as e:
        print(f"      API call failed: {str(e)}")
        return []
    except Exception as e:
        print(f"      Error processing API response: {str(e)}")
        return []

def format_match_data_for_excel(match):
    """
    Format match data for Excel display (same as before)
    """
    try:
        from datetime import datetime, timedelta
        
        # Format date from YYYY-MM-DD to DD/MM/YYYY
        if match['matchTime']:
            try:
                date_obj = datetime.strptime(match['matchTime'], '%Y-%m-%d %H:%M:%S')
                match['date_formatted'] = date_obj.strftime('%d/%m/%Y')
            except:
                match['date_formatted'] = ''
        else:
            match['date_formatted'] = ''
        
        # Format local time to HH:MM
        if match['matchTime']:
            try:
                time_obj = datetime.strptime(match['matchTime'], '%Y-%m-%d %H:%M:%S')
                match['time_local_formatted'] = time_obj.strftime('%H:%M')
            except:
                match['time_local_formatted'] = ''
        else:
            match['time_local_formatted'] = ''
        
        # Format UTC time to HH:MM
        if match['matchTimeUTC']:
            try:
                utc_time_obj = datetime.strptime(match['matchTimeUTC'], '%Y-%m-%d %H:%M:%S')
                match['time_utc_formatted'] = utc_time_obj.strftime('%H:%M')
            except:
                match['time_utc_formatted'] = ''
        else:
            match['time_utc_formatted'] = ''
        
        # Convert UTC time to Tallinn time (UTC+2)
        if match['matchTimeUTC']:
            try:
                utc_time_obj = datetime.strptime(match['matchTimeUTC'], '%Y-%m-%d %H:%M:%S')
                tallinn_time = utc_time_obj + timedelta(hours=2)
                match['time_tallinn_formatted'] = tallinn_time.strftime('%H:%M')
            except:
                match['time_tallinn_formatted'] = ''
        else:
            match['time_tallinn_formatted'] = ''
        
        # Convert UTC time to Medellin time (UTC-5)
        if match['matchTimeUTC']:
            try:
                utc_time_obj = datetime.strptime(match['matchTimeUTC'], '%Y-%m-%d %H:%M:%S')
                medellin_time = utc_time_obj - timedelta(hours=5)
                match['time_medellin_formatted'] = medellin_time.strftime('%H:%M')
            except:
                match['time_medellin_formatted'] = ''
        else:
            match['time_medellin_formatted'] = ''
        
        # Create game description
        if match['competitor1'] and match['competitor2']:
            match['game'] = f"{match['competitor1']} vs {match['competitor2']}"
        else:
            match['game'] = 'TBD'
        
        return match
        
    except Exception as e:
        print(f"Error formatting match data: {str(e)}")
        return match

def fetch_matches_for_competitions(api_key, competitions):
    """
    Fetch matches for all competitions
    """
    print(f"\n=== FETCHING MATCHES FOR ALL COMPETITIONS ===")
    
    competitions_with_matches = []
    
    for comp_idx, competition in enumerate(competitions):
        print(f"\nProcessing competition {comp_idx + 1}/{len(competitions)}: {competition['competitionName']}")
        
        matches = fetch_matches_for_competition(api_key, competition['competitionId'], competition['competitionName'])
        competition['matches'] = matches
        
        competitions_with_matches.append(competition)
        print(f"  Added {len(matches)} matches to competition")
    
    return competitions_with_matches

def create_excel_file_with_competitions(competitions, output_path):
    """
    Create a nicely formatted Excel file with all the match data
    Each competition gets its own table with all the validation checks
    We also preserve any manual notes/checks that were added to the previous file
    """
    print(f"\n=== CREATING EXCEL FILE WITH COMPETITIONS ===")
    print(f"Output path: {output_path}")
    
    try:
        # Load the previous Excel file if it exists
        # This way we keep any manual notes or checks that someone added
        existing_matches_map = load_existing_matches(output_path)
        if existing_matches_map:
            print("Loaded existing matches from previous file to preserve manual checks.")
        
        # Load state so we don't re-add matches the user has deleted
        state = load_fetch_state(output_path)
        last_written = state.get("last_written", {})
        deleted = state.get("deleted", {})
        # Only detect "newly deleted" when the Excel file actually existed.
        # If the user deleted the file and re-runs, we must not treat all matches as deleted.
        output_exists = Path(output_path).exists()
        if output_exists and existing_matches_map:
            for comp_id_str, written_ids in last_written.items():
                current_in_file = set(existing_matches_map.get(comp_id_str, {}).keys())
                written_set = set(written_ids) if isinstance(written_ids, list) else set()
                newly_deleted = written_set - current_in_file
                # Only treat as "user deleted" when a small number of rows were removed.
                # If the majority of matches are missing (e.g. file was recreated or replaced),
                # clear deleted for this comp so we re-add all matches and restore full count.
                if newly_deleted and written_set:
                    removed_ratio = len(newly_deleted) / len(written_set)
                    if removed_ratio <= 0.5:
                        deleted[comp_id_str] = list(set(deleted.get(comp_id_str, [])) | newly_deleted)
                    else:
                        deleted[comp_id_str] = []  # clear so missing matches are re-added
        else:
            # No existing file: start fresh, do not treat any match as user-deleted
            deleted = {}
        state["deleted"] = deleted
        
        # Create new workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create "FMM automation" sheet (same name as before)
        ws = wb.create_sheet("FMM automation")
        print("Created FMM automation sheet")
        
        current_row = 1
        total_new_matches = 0  # matches added since last launch (not in previous last_written)

        for competition_idx, competition in enumerate(competitions):
            print(f"\nProcessing competition {competition_idx + 1}/{len(competitions)}: {competition['competitionName']}")
            
            # Build dynamic headers (Coretools/WHST depend on LiveDataSource)
            live_source = (competition.get('liveDataSource') or '').strip()
            is_meb = live_source.lower() in ['match events bridge', 'isd', 'match events']
            coretools_header = 'No need' if is_meb else 'Pre-game Coretools Mapping check'
            whst_header = 'Pre-game WHST Live Data Source ISD' if is_meb else 'Pre-game WHST Live Data Source GS Live Stats'

            headers = [
                'League',
                'Date',
                'Time Local',
                'Time UTC',
                'Time Tallinn',
                'Time Medellin',
                'Game ID',
                'Game',
                'Pre-game DM check',
                coretools_header,
                whst_header,
                'Live game Statistician check',
                'Live game Webcast check',
                'End game Past match data'
            ]
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=current_row, column=col_idx, value=header)
            current_row += 1
            
            # Add league and competition info row
            league_comp_info_text = f"{competition['leagueName']} (League ID {competition['leagueId']}) - {competition['competitionName']} (Competition ID {competition['competitionId']})"
            ws.cell(row=current_row, column=1, value=league_comp_info_text)
            # Merge cells for the league/competition info row across all columns
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
            current_row += 1
            
            # Add matches for this competition (don't re-add user-deleted matches)
            matches = competition.get('matches', [])
            comp_id_str = str(competition.get('competitionId', ''))
            existing_comp_matches = existing_matches_map.get(comp_id_str, {})
            deleted_for_comp = deleted.get(comp_id_str, [])
            # If API returned many more matches than the file has, treat file as out of date and include all
            if matches and len(existing_comp_matches) > 0 and len(matches) >= 2 * len(existing_comp_matches):
                deleted_for_comp = []
            merged_matches = merge_matches_with_existing(matches, existing_comp_matches, deleted_match_ids=deleted_for_comp)
            # If using the deleted list would hide a large share of API matches, ignore it so all found matches are included
            if matches and deleted_for_comp and len(merged_matches) < len(matches) * 0.95:
                merged_matches = merge_matches_with_existing(matches, existing_comp_matches, deleted_match_ids=[])
                state["deleted"][comp_id_str] = []  # clear so next run we don't exclude them again
            competition['matches'] = merged_matches
            # Count new matches (not in previous run's written set)
            prev_written = {str(x).strip() for x in (last_written.get(comp_id_str, []) or [])}
            total_new_matches += sum(1 for m in merged_matches if str(m.get('matchId', '')).strip() not in prev_written)
            # Remember what we wrote so next run we can detect newly deleted rows
            state["last_written"][comp_id_str] = [str(m.get('matchId', '')) for m in merged_matches if m.get('matchId')]

            if merged_matches:
                new_match_ids = {str(match.get('matchId', '')).strip() for match in matches if match.get('matchId')}
                total_matches = len(merged_matches)
                preserved_matches = total_matches - len(new_match_ids)
                print(f"    Adding {total_matches} matches to table (new/updated: {len(new_match_ids)}, carried over: {max(preserved_matches, 0)})")

                for match in merged_matches:
                    ws.cell(row=current_row, column=1, value=match.get('league_column_note', ''))  # Column A (League) - preserve user notes
                    ws.cell(row=current_row, column=2, value=match.get('date_formatted', ''))  # Date
                    ws.cell(row=current_row, column=3, value=match.get('time_local_formatted', ''))  # Time Local
                    ws.cell(row=current_row, column=4, value=match.get('time_utc_formatted', ''))  # Time UTC
                    ws.cell(row=current_row, column=5, value=match.get('time_tallinn_formatted', ''))  # Time Tallinn
                    ws.cell(row=current_row, column=6, value=match.get('time_medellin_formatted', ''))  # Time Medellin
                    ws.cell(row=current_row, column=7, value=match.get('matchId', ''))  # Game ID
                    ws.cell(row=current_row, column=8, value=match.get('game', ''))  # Game
                    ws.cell(row=current_row, column=9, value=match.get('livestream_status', 'N/A'))  # Pre-game DM check
                    ws.cell(row=current_row, column=10, value=match.get('coretools_check', ''))  # Pre-game Coretools Mapping check / No need
                    ws.cell(row=current_row, column=11, value=match.get('whst_live_data_source_match', 'N/A'))  # Pre-game WHST Live Data Source
                    ws.cell(row=current_row, column=12, value=match.get('publish_connection_status', 'N/A'))  # Live game Statistician check
                    ws.cell(row=current_row, column=13, value=match.get('webcast_status', ''))  # Live game Webcast check
                    # End game Past match data: "Yes"=ok, "Match check required"=issues (with HS link), "N/A - Match check required"=check not run, "Too early"=not yet played.
                    end_game_cell = ws.cell(row=current_row, column=14, value=match.get('end_game_status', ''))
                    if match.get('end_game_hs_url'):
                        end_game_cell.hyperlink = match.get('end_game_hs_url')
                        from openpyxl.styles import Font
                        end_game_cell.font = Font(underline="single", color="0563C1", size=10)
                    # Future checklist columns (initialize empty if more headers are added later)
                    for col_idx in range(15, len(headers) + 1):
                        ws.cell(row=current_row, column=col_idx, value='')
                    current_row += 1
            else:
                print(f"    No matches found for this competition")
            
            # Add "Last check" row with today's date (merged across all columns)
            today = datetime.now().strftime('%d/%m')
            ws.cell(row=current_row, column=1, value=f"Last check {today}")
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
            current_row += 1
            
            # Add empty row gap between tables (except for last table)
            if competition_idx < len(competitions) - 1:
                current_row += 1
        
        # Apply table styling
        print("Applying table styling...")
        apply_table_styling(ws, len(competitions))
        
        # Save the workbook
        wb.save(output_path)
        save_fetch_state(output_path, state)
        print(f"Excel file created successfully: {output_path}")

        # Slack: send summary with UTC timestamp, competitions, total matches, new matches, and any check issues
        total_matches = sum(len(c.get('matches', [])) for c in competitions)
        utc_ts = datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')
        slack_text = (
            f"FMM run completed at {utc_ts}\n"
            f"• Competitions processed: {len(competitions)}\n"
            f"• Total matches in Excel: {total_matches}\n"
            f"• New matches since last launch: {total_new_matches}"
        )
        issues = collect_check_issues(competitions)
        if issues:
            slack_text += "\n\n*Issues found:*"
            for comp_issue in issues:
                slack_text += (
                    f"\n\n*{comp_issue['league_name']}* (League ID {comp_issue['league_id']}) — "
                    f"*{comp_issue['competition_name']}* (Competition ID {comp_issue['competition_id']})"
                )
                for m in comp_issue["matches"]:
                    checks_str = ", ".join(m["failed_checks"])
                    slack_text += f"\n  • {m['game']} (ID {m['matchId']}): {checks_str}"
                    if m.get("webcast_url"):
                        slack_text += f" | Webcast: <{m['webcast_url']}>"
                    if m.get("hs_url"):
                        slack_text += f" | HS: <{m['hs_url']}>"
        if send_slack_message(slack_text):
            print("Slack notification sent to #notifications-fmm")
        else:
            print("Slack not sent (set SLACK_BOT_TOKEN to enable)")

        return True
        
    except Exception as e:
        print(f"Error creating Excel file: {str(e)}")
        return False

def apply_table_styling(ws, num_tables):
    """
    Apply styling to the FMM automation sheet (same as before)
    """
    try:
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        # Define colors for different row types
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")  # Dark blue
        league_info_fill = PatternFill(start_color="8DB4E2", end_color="8DB4E2", fill_type="solid")  # Light blue
        last_check_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")  # Very light blue
        
        # Define fonts
        header_font = Font(color="FFFFFF", bold=True, size=11)  # White bold text for headers
        normal_font = Font(size=10)
        bold_font = Font(bold=True, size=10)
        
        # Define alignment
        center_alignment = Alignment(horizontal="center", vertical="center")
        left_alignment = Alignment(horizontal="left", vertical="center")
        
        # Define border
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Set column widths (extended for new checklist columns)
        column_widths = {
            'A': 12,  # League
            'B': 12,  # Date
            'C': 12,  # Time Local
            'D': 12,  # Time UTC
            'E': 12,  # Time Tallinn
            'F': 12,  # Time Medellin
            'G': 12,  # Game ID
            'H': 60,  # Game (twice as wide for team names)
            'I': 20,  # Pre-game DM check
            'J': 35,  # Pre-game Coretools Mapping check/No need
            'K': 38,  # Pre-game WHST Live Data Source (ISD/GS Live Stats)
            'L': 28,  # Live game Statistician check
            'M': 25,  # Live game Webcast check
            'N': 28   # End game Past match data
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Track current row and apply styling
        current_row = 1
        
        for table_idx in range(num_tables):
            # Find the end of this table by looking for "Last check"
            table_end_row = current_row
            for row in range(current_row, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=1).value
                if cell_value and "Last check" in str(cell_value):
                    table_end_row = row
                    break
            
            # Apply styling to this table
            last_col = ws.max_column
            for row in range(current_row, table_end_row + 1):
                # Apply borders to all cells
                for col in range(1, last_col + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
                    cell.alignment = left_alignment
                    cell.font = normal_font
                
                # Determine row type and apply specific styling
                cell_a = ws.cell(row=row, column=1).value
                
                if cell_a and str(cell_a).strip() in ['League']:
                    # Header row
                    for col in range(1, last_col + 1):
                        cell = ws.cell(row=row, column=col)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = center_alignment
                
                elif cell_a and ("League ID" in str(cell_a) or "League" in str(cell_a)) and row > current_row:
                    # League info row
                    for col in range(1, 9):
                        cell = ws.cell(row=row, column=col)
                        cell.fill = league_info_fill
                        cell.font = bold_font
                
                elif cell_a and "Last check" in str(cell_a):
                    # Last check row (merged from A to H)
                    for col in range(1, 9):
                        cell = ws.cell(row=row, column=col)
                        cell.fill = last_check_fill
                        cell.font = bold_font
                        cell.alignment = center_alignment
                
                else:
                    # Check if this is a match row (has a Game ID in column 7)
                    game_id = ws.cell(row=row, column=7).value
                    if game_id:
                        # This is a match row - center-align columns 9, 11, and 12
                        ws.cell(row=row, column=9).alignment = center_alignment  # Pre-game DM check
                        ws.cell(row=row, column=11).alignment = center_alignment  # Pre-game WHST Live Data Source
                        ws.cell(row=row, column=12).alignment = center_alignment  # Live game Statistician check
            
            # Move to next table (skip empty gap row)
            current_row = table_end_row + 2
        
        print("✓ Table styling applied successfully")
        
    except Exception as e:
        print(f"Error applying table styling: {str(e)}")

def main():
    """
    This is where everything starts
    It asks for an API key, loads the list of competitions to check,
    fetches all the match data, runs all the validation checks,
    and creates an Excel file with everything
    """
    # Ask the user for their API key so we can make API calls
    api_key = get_user_api_key()
    if not api_key:
        return
    
    # Load the list of competitions we want to process
    # This is stored in a JSON file so you can easily add/remove competitions
    whitelist_ids, whitelist_config = load_competition_whitelist()
    
    if not whitelist_ids:
        print("No competitions in whitelist. Please add competitions to competition_whitelist.json")
        show_whitelist_management_info(whitelist_config)
        return
    
    print(f"\nSuccessfully loaded whitelist with {len(whitelist_ids)} competitions")
    
    # Process only the competitions in our whitelist
    # This is faster than fetching all competitions and then filtering
    print(f"\n=== PROCESSING WHITELISTED COMPETITIONS ===")
    competitions_with_matches = process_whitelisted_competitions_directly(api_key, whitelist_ids, whitelist_config)
    
    if not competitions_with_matches:
        print("No whitelisted competitions found or no matches available.")
        return
    
    # Create output file path that sits next to the script
    script_dir = Path(__file__).parent
    output_file = script_dir / "football_competitions_fetch.xlsx"
    
    # Create Excel file with competitions and matches
    success = create_excel_file_with_competitions(competitions_with_matches, str(output_file))
    
    if success:
        print(f"\n✅ SUCCESS! Created Excel file with {len(competitions_with_matches)} competition tables")
        print(f"File saved to: {output_file}")
        
        # Show summary
        total_matches = 0
        for comp in competitions_with_matches:
            total_matches += len(comp.get('matches', []))
        
        print(f"\n📊 Summary:")
        print(f"  - Competitions in whitelist: {len(whitelist_ids)}")
        print(f"  - Competitions processed: {len(competitions_with_matches)}")
        print(f"  - Total matches: {total_matches}")
        
        print(f"\n💡 To add more competitions, edit: competition_whitelist.json")
    else:
        print("\n❌ Failed to create Excel file")

if __name__ == "__main__":
    main()
