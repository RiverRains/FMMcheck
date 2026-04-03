import re
import logging
from collections import defaultdict
from datetime import datetime, timedelta, time, timezone
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, NamedStyle

from config.settings import should_notify_resolved_issues, get_notification_state_retention_days
from storage.notification_state import NotificationStateManager, prune_resolved_issues
from storage.state import StateManager
from processing.match_evaluator import collect_check_issues, flatten_check_issues
from notifications.slack import send_slack_message

logger = logging.getLogger(__name__)

def parse_date_string(date_value):
    if not date_value:
        return None
    if isinstance(date_value, datetime):
        return date_value.date()
    date_str = str(date_value).strip()
    if not date_str:
        return None
    date_formats = ['%d/%m/%Y', '%d/%m/%y', '%Y-%m-%d', '%Y/%m/%d', '%m/%d/%Y', '%m/%d/%y']
    for fmt in date_formats:
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    return None

def parse_time_string(time_value):
    if not time_value:
        return None
    if isinstance(time_value, datetime):
        return time_value.time()
    if isinstance(time_value, time):
        return time_value
    time_str = str(time_value).strip()
    if not time_str:
        return None
    time_formats = ['%H:%M:%S', '%H:%M']
    for fmt in time_formats:
        try:
            return datetime.strptime(time_str, fmt).time()
        except ValueError:
            continue
    return None

def match_sort_key(match):
    date_obj = parse_date_string(match.get('date_formatted'))
    time_obj = parse_time_string(match.get('time_local_formatted'))
    if date_obj is None:
        date_obj = datetime.max.date()
    if time_obj is None:
        time_obj = time.min
    match_id = str(match.get('matchId', ''))
    return (date_obj, time_obj, match_id)

def load_existing_matches(output_path):
    existing_matches = {}
    target_path = Path(output_path)

    if not target_path.exists():
        return existing_matches

    try:
        wb = openpyxl.load_workbook(output_path, read_only=True)
    except Exception as e:
        logger.warning(f"Could not load existing Excel file ({e}). Continuing without previous data.")
        return existing_matches

    if "FMM automation" not in wb.sheetnames:
        wb.close()
        return existing_matches

    ws = wb["FMM automation"]
    today = datetime.now().date()
    two_weeks_ago = today - timedelta(days=14)
    current_competition_id = None

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        cell_a = row[0]
        if cell_a == 'League':
            continue

        if cell_a and "Competition ID" in str(cell_a):
            comp_match = re.search(r"Competition ID\s+(\d+)", str(cell_a))
            current_competition_id = comp_match.group(1) if comp_match else None
            if current_competition_id:
                existing_matches.setdefault(current_competition_id, {})
            continue

        if cell_a and isinstance(cell_a, str) and cell_a.startswith("Last check"):
            current_competition_id = None
            continue

        if current_competition_id:
            match_id_cell = row[6]
            if match_id_cell:
                match_id = str(match_id_cell).strip()

                league_cell = row[0]
                league_column_note = str(league_cell).strip() if league_cell else ''

                date_cell = row[1]
                if isinstance(date_cell, datetime):
                    date_formatted = date_cell.strftime('%d/%m/%Y')
                else:
                    date_formatted = str(date_cell or '').strip()

                time_local_cell = row[2]
                if isinstance(time_local_cell, datetime) or isinstance(time_local_cell, time):
                    time_local_formatted = time_local_cell.strftime('%H:%M')
                else:
                    time_local_formatted = str(time_local_cell or '').strip()

                time_utc_cell = row[3]
                if isinstance(time_utc_cell, datetime) or isinstance(time_utc_cell, time):
                    time_utc_formatted = time_utc_cell.strftime('%H:%M')
                else:
                    time_utc_formatted = str(time_utc_cell or '').strip()

                time_tallinn_cell = row[4]
                if isinstance(time_tallinn_cell, datetime) or isinstance(time_tallinn_cell, time):
                    time_tallinn_formatted = time_tallinn_cell.strftime('%H:%M')
                else:
                    time_tallinn_formatted = str(time_tallinn_cell or '').strip()

                time_medellin_cell = row[5]
                if isinstance(time_medellin_cell, datetime) or isinstance(time_medellin_cell, time):
                    time_medellin_formatted = time_medellin_cell.strftime('%H:%M')
                else:
                    time_medellin_formatted = str(time_medellin_cell or '').strip()

                game_value = row[7]
                livestream_status = str(row[8] or '').strip()
                coretools_check = str(row[9] or '').strip()
                whst_status = str(row[10] or '').strip()
                publish_status = str(row[11] or '').strip()
                webcast_status = str(row[12] or '').strip()
                end_game_status = str(row[13] or '').strip()

                match_date_obj = parse_date_string(date_formatted)
                if match_date_obj and match_date_obj < two_weeks_ago:
                    logger.info(f"Removing match {match_id} (older than 2 weeks, date: {date_formatted})")
                    continue

                existing_matches[current_competition_id][match_id] = {
                    'matchId': match_id,
                    'league_column_note': league_column_note,
                    'date_formatted': date_formatted,
                    'time_local_formatted': time_local_formatted,
                    'time_utc_formatted': time_utc_formatted,
                    'time_tallinn_formatted': time_tallinn_formatted,
                    'time_medellin_formatted': time_medellin_formatted,
                    'game': str(game_value or '').strip(),
                    'livestream_status': livestream_status or 'N/A',
                    'coretools_check': coretools_check,
                    'whst_live_data_source_match': whst_status or 'N/A',
                    'publish_connection_status': publish_status or 'N/A',
                    'webcast_status': webcast_status,
                    'end_game_status': end_game_status
                }
    wb.close()
    return existing_matches

def merge_matches_with_existing(new_matches, existing_matches, deleted_match_ids=None):
    merged_matches_map = {str(mid): existing.copy() for mid, existing in existing_matches.items()}
    deleted_set = {str(mid).strip() for mid in (deleted_match_ids or [])}

    for match in new_matches:
        match_id = str(match.get('matchId', '')).strip()
        if not match_id:
            continue
        if match_id in deleted_set and match_id not in existing_matches:
            continue

        existing_entry = existing_matches.get(match_id)
        merged_entry = existing_entry.copy() if existing_entry else {}
        merged_entry.update(match)

        if existing_entry:
            merged_entry['coretools_check'] = existing_entry.get('coretools_check', '')
            merged_entry['league_column_note'] = existing_entry.get('league_column_note', '')
            # Preserve previous statistician/webcast results when new run has no real value
            new_pub = match.get('publish_connection_status', '')
            if not new_pub or new_pub == 'N/A':
                merged_entry['publish_connection_status'] = existing_entry.get('publish_connection_status', 'N/A')
            new_wc = match.get('webcast_status', '')
            if not new_wc or new_wc == 'N/A':
                merged_entry['webcast_status'] = existing_entry.get('webcast_status', '')
        else:
            merged_entry.setdefault('coretools_check', '')
            merged_entry.setdefault('league_column_note', '')
            merged_entry.setdefault('end_game_status', match.get('end_game_status', ''))
            merged_entry.setdefault('webcast_status', match.get('webcast_status', ''))

        merged_entry.setdefault('livestream_status', match.get('livestream_status', 'N/A'))
        merged_entry.setdefault('whst_live_data_source_match', match.get('whst_live_data_source_match', 'N/A'))
        merged_entry.setdefault('publish_connection_status', match.get('publish_connection_status', 'N/A'))
        merged_entry.setdefault('time_utc_formatted', match.get('time_utc_formatted', ''))
        merged_entry.setdefault('time_tallinn_formatted', match.get('time_tallinn_formatted', ''))
        merged_entry.setdefault('time_medellin_formatted', match.get('time_medellin_formatted', ''))
        merged_entry.setdefault('game', match.get('game', ''))
        merged_entry.setdefault('date_formatted', match.get('date_formatted', ''))
        merged_entry.setdefault('time_local_formatted', match.get('time_local_formatted', ''))

        merged_matches_map[match_id] = merged_entry

    merged_list = list(merged_matches_map.values())
    merged_list.sort(key=match_sort_key)
    return merged_list

def apply_table_styling(wb, ws, num_tables):
    try:
        # Create NamedStyles
        if "fmm_header" not in wb.named_styles:
            header_style = NamedStyle(name="fmm_header")
            header_style.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_style.font = Font(color="FFFFFF", bold=True, size=11)
            header_style.alignment = Alignment(horizontal="center", vertical="center")
            header_style.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            wb.add_named_style(header_style)
            
        if "fmm_league_info" not in wb.named_styles:
            league_info_style = NamedStyle(name="fmm_league_info")
            league_info_style.fill = PatternFill(start_color="8DB4E2", end_color="8DB4E2", fill_type="solid")
            league_info_style.font = Font(bold=True, size=10)
            league_info_style.alignment = Alignment(horizontal="left", vertical="center")
            league_info_style.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            wb.add_named_style(league_info_style)
            
        if "fmm_last_check" not in wb.named_styles:
            last_check_style = NamedStyle(name="fmm_last_check")
            last_check_style.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
            last_check_style.font = Font(bold=True, size=10)
            last_check_style.alignment = Alignment(horizontal="center", vertical="center")
            last_check_style.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            wb.add_named_style(last_check_style)
            
        if "fmm_normal" not in wb.named_styles:
            normal_style = NamedStyle(name="fmm_normal")
            normal_style.font = Font(size=10)
            normal_style.alignment = Alignment(horizontal="left", vertical="center")
            normal_style.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            wb.add_named_style(normal_style)
        
        column_widths = {
            'A': 12, 'B': 12, 'C': 12, 'D': 12, 'E': 12, 'F': 12, 'G': 12,
            'H': 60, 'I': 20, 'J': 35, 'K': 38, 'L': 28, 'M': 25, 'N': 28
        }
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
            
        current_row = 1
        for table_idx in range(num_tables):
            table_end_row = current_row
            for row in range(current_row, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=1).value
                if cell_value and "Last check" in str(cell_value):
                    table_end_row = row
                    break
            
            last_col = ws.max_column
            for row in range(current_row, table_end_row + 1):
                cell_a = ws.cell(row=row, column=1).value
                
                if cell_a and str(cell_a).strip() in ['League']:
                    for col in range(1, last_col + 1):
                        ws.cell(row=row, column=col).style = "fmm_header"
                elif cell_a and ("League ID" in str(cell_a) or "League" in str(cell_a)) and row > current_row:
                    for col in range(1, last_col + 1):
                        cell = ws.cell(row=row, column=col)
                        if col <= 8:
                            cell.style = "fmm_league_info"
                        else:
                            cell.style = "fmm_normal"
                elif cell_a and "Last check" in str(cell_a):
                    for col in range(1, last_col + 1):
                        cell = ws.cell(row=row, column=col)
                        if col <= 8:
                            cell.style = "fmm_last_check"
                        else:
                            cell.style = "fmm_normal"
                else:
                    for col in range(1, last_col + 1):
                        ws.cell(row=row, column=col).style = "fmm_normal"
                    game_id = ws.cell(row=row, column=7).value
                    if game_id:
                        ws.cell(row=row, column=9).alignment = Alignment(horizontal="center", vertical="center")
                        ws.cell(row=row, column=11).alignment = Alignment(horizontal="center", vertical="center")
                        ws.cell(row=row, column=12).alignment = Alignment(horizontal="center", vertical="center")
            
            current_row = table_end_row + 2
        logger.info("Table styling applied successfully using NamedStyles")
    except Exception as e:
        logger.error(f"Error applying table styling: {str(e)}")


def _utc_now_iso():
    return datetime.now(timezone.utc).isoformat()


def _sorted_issue_records(issue_records):
    return sorted(
        issue_records,
        key=lambda issue: (
            issue.get("league_name", ""),
            issue.get("competition_name", ""),
            issue.get("game", ""),
            issue.get("match_id", ""),
            issue.get("check_name", ""),
        ),
    )


def _build_open_issue_state(current_issue_map, previous_open_issues, now_iso):
    open_issue_state = {}

    for issue_key, issue in current_issue_map.items():
        previous_issue = previous_open_issues.get(issue_key, {})
        open_issue_state[issue_key] = {
            **issue,
            "first_seen_at": previous_issue.get("first_seen_at", now_iso),
            "last_seen_at": now_iso,
        }

    return open_issue_state


def _build_resolved_issue_state(previous_resolved_issues, previous_open_issues, current_issue_map, resolved_issue_keys, now_iso):
    resolved_issue_state = {
        issue_key: issue
        for issue_key, issue in (previous_resolved_issues or {}).items()
        if issue_key not in current_issue_map
    }

    for issue_key in resolved_issue_keys:
        previous_issue = previous_open_issues.get(issue_key)
        if not previous_issue:
            continue

        resolved_issue_state[issue_key] = {
            **previous_issue,
            "resolved_at": now_iso,
        }

    return resolved_issue_state


def _append_issue_section(lines, title, issue_records):
    if not issue_records:
        return

    grouped_records = defaultdict(list)
    for issue in _sorted_issue_records(issue_records):
        competition_key = (
            issue.get("league_name", ""),
            issue.get("league_id", ""),
            issue.get("competition_name", ""),
            issue.get("competition_id", ""),
        )
        grouped_records[competition_key].append(issue)

    lines.append("")
    lines.append(f"*{title}:*")

    for (league_name, league_id, competition_name, competition_id), records in grouped_records.items():
        lines.append("")
        lines.append(
            f"*{league_name}* (League ID {league_id}) — *{competition_name}* (Competition ID {competition_id})"
        )
        for issue in records:
            line = f"  • {issue.get('game', 'Match')} (ID {issue.get('match_id', '')}): {issue.get('check_name', '')}"
            detail_url = issue.get("detail_url", "")
            detail_label = issue.get("detail_label", "")
            if detail_url and detail_label:
                line += f" | {detail_label}: <{detail_url}>"
            lines.append(line)


def _build_slack_summary_text(competitions, total_matches, total_new_matches, total_open_issues, new_issue_records, resolved_issue_records):
    notify_resolved = should_notify_resolved_issues()
    utc_ts = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    lines = [
        f"FMM run completed at {utc_ts}",
        f"• Competitions processed: {len(competitions)}",
        f"• Total matches in Excel: {total_matches}",
        f"• New matches added: {total_new_matches}",
        f"• Open issues: {total_open_issues}",
        f"• New issues: {len(new_issue_records)}",
    ]

    if notify_resolved:
        lines.append(f"• Resolved issues: {len(resolved_issue_records)}")

    if not new_issue_records and not (notify_resolved and resolved_issue_records):
        lines.append("\nNo new issues detected.")

    _append_issue_section(lines, "New issues", new_issue_records)
    if notify_resolved:
        _append_issue_section(lines, "Resolved issues", resolved_issue_records)

    return "\n".join(lines)

def create_excel_file_with_competitions(competitions, output_path, whitelist_config=None):
    logger.info(f"Creating Excel file with competitions at: {output_path}")
    try:
        existing_matches_map = load_existing_matches(output_path)
        if existing_matches_map:
            logger.info("Loaded existing matches from previous file to preserve manual checks.")

        state_mgr = StateManager(output_path)
        state = state_mgr.load_fetch_state()
        notification_state_mgr = NotificationStateManager(output_path)
        notification_state = notification_state_mgr.load_state()
        notification_state_mgr.ensure_state_file()
        last_written = state.get("last_written", {})
        deleted = state.get("deleted", {})
        
        output_exists = Path(output_path).exists()
        if output_exists and existing_matches_map:
            for comp_id_str, written_ids in last_written.items():
                current_in_file = set(existing_matches_map.get(comp_id_str, {}).keys())
                written_set = set(written_ids) if isinstance(written_ids, list) else set()
                newly_deleted = written_set - current_in_file
                if newly_deleted and written_set:
                    removed_ratio = len(newly_deleted) / len(written_set)
                    if removed_ratio <= 0.5:
                        deleted[comp_id_str] = list(set(deleted.get(comp_id_str, [])) | newly_deleted)
                    else:
                        deleted[comp_id_str] = []
        else:
            deleted = {}
        state["deleted"] = deleted
        
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("FMM automation")
        
        current_row = 1
        total_new_matches = 0

        for competition_idx, competition in enumerate(competitions):
            logger.info(f"Processing competition {competition_idx + 1}/{len(competitions)}: {competition['competitionName']}")
            live_source = (competition.get('liveDataSource') or '').strip()
            is_meb = live_source.lower() in ['match events bridge', 'isd', 'match events']
            coretools_header = 'No need' if is_meb else 'Pre-game Coretools Mapping check'
            whst_header = 'Pre-game WHST Live Data Source ISD' if is_meb else 'Pre-game WHST Live Data Source GS Live Stats'

            headers = [
                'League', 'Date', 'Time Local', 'Time UTC', 'Time Tallinn', 'Time Medellin',
                'Game ID', 'Game', 'Pre-game DM check', coretools_header, whst_header,
                'Live game Statistician check', 'Live game Webcast check', 'End game Past match data'
            ]
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=current_row, column=col_idx, value=header)
            current_row += 1
            
            league_comp_info_text = f"{competition['leagueName']} (League ID {competition['leagueId']}) - {competition['competitionName']} (Competition ID {competition['competitionId']})"
            ws.cell(row=current_row, column=1, value=league_comp_info_text)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
            current_row += 1
            
            matches = competition.get('matches', [])
            comp_id_str = str(competition.get('competitionId', ''))
            existing_comp_matches = existing_matches_map.get(comp_id_str, {})
            deleted_for_comp = deleted.get(comp_id_str, [])
            
            if matches and len(existing_comp_matches) > 0 and len(matches) >= 2 * len(existing_comp_matches):
                deleted_for_comp = []
            merged_matches = merge_matches_with_existing(matches, existing_comp_matches, deleted_match_ids=deleted_for_comp)
            
            if matches and deleted_for_comp and len(merged_matches) < len(matches) * 0.95:
                merged_matches = merge_matches_with_existing(matches, existing_comp_matches, deleted_match_ids=[])
                state["deleted"][comp_id_str] = []
            competition['matches'] = merged_matches
            
            prev_written = {str(x).strip() for x in (last_written.get(comp_id_str, []) or [])}
            total_new_matches += sum(1 for m in merged_matches if str(m.get('matchId', '')).strip() not in prev_written)
            state["last_written"][comp_id_str] = [str(m.get('matchId', '')) for m in merged_matches if m.get('matchId')]

            if merged_matches:
                new_match_ids = {str(match.get('matchId', '')).strip() for match in matches if match.get('matchId')}
                total_matches = len(merged_matches)
                preserved_matches = total_matches - len(new_match_ids)
                logger.info(f"Adding {total_matches} matches to table (new/updated: {len(new_match_ids)}, carried over: {max(preserved_matches, 0)})")

                for match in merged_matches:
                    ws.cell(row=current_row, column=1, value=match.get('league_column_note', ''))
                    ws.cell(row=current_row, column=2, value=match.get('date_formatted', ''))
                    ws.cell(row=current_row, column=3, value=match.get('time_local_formatted', ''))
                    ws.cell(row=current_row, column=4, value=match.get('time_utc_formatted', ''))
                    ws.cell(row=current_row, column=5, value=match.get('time_tallinn_formatted', ''))
                    ws.cell(row=current_row, column=6, value=match.get('time_medellin_formatted', ''))
                    ws.cell(row=current_row, column=7, value=match.get('matchId', ''))
                    ws.cell(row=current_row, column=8, value=match.get('game', ''))
                    ws.cell(row=current_row, column=9, value=match.get('livestream_status', 'N/A'))
                    ws.cell(row=current_row, column=10, value=match.get('coretools_check', ''))
                    ws.cell(row=current_row, column=11, value=match.get('whst_live_data_source_match', 'N/A'))
                    ws.cell(row=current_row, column=12, value=match.get('publish_connection_status', 'N/A'))
                    ws.cell(row=current_row, column=13, value=match.get('webcast_status', ''))
                    
                    end_game_cell = ws.cell(row=current_row, column=14, value=match.get('end_game_status', ''))
                    if match.get('end_game_hs_url'):
                        end_game_cell.hyperlink = match.get('end_game_hs_url')
                        end_game_cell.font = Font(underline="single", color="0563C1", size=10)
                    for col_idx in range(15, len(headers) + 1):
                        ws.cell(row=current_row, column=col_idx, value='')
                    current_row += 1
            else:
                logger.info("No matches found for this competition")
            
            today = datetime.now().strftime('%d/%m')
            ws.cell(row=current_row, column=1, value=f"Last check {today}")
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
            current_row += 1
            
            if competition_idx < len(competitions) - 1:
                current_row += 1
        
        apply_table_styling(wb, ws, len(competitions))

        # Add whitelist tab if config provided
        if whitelist_config:
            write_whitelist_sheet(wb, whitelist_config)

        # Ensure the directory exists
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        state_mgr.save_fetch_state(state)
        logger.info(f"Excel file created successfully: {output_path}")

        previous_open_issues = notification_state.get("open_issues", {})
        previous_resolved_issues = notification_state.get("resolved_issues", {})
        now_iso = _utc_now_iso()

        logger.info(f"Notification state loaded: {len(previous_open_issues)} previous open issues, {len(previous_resolved_issues)} previously resolved")

        total_matches = sum(len(c.get('matches', [])) for c in competitions)
        issue_groups = collect_check_issues(competitions)
        current_issue_map = flatten_check_issues(issue_groups)
        current_issue_keys = set(current_issue_map)
        previous_issue_keys = set(previous_open_issues)
        new_issue_keys = current_issue_keys - previous_issue_keys
        resolved_issue_keys = previous_issue_keys - current_issue_keys

        logger.info(f"Issue diff: {len(current_issue_keys)} current, {len(previous_issue_keys)} previous, {len(new_issue_keys)} new, {len(resolved_issue_keys)} resolved")

        next_open_issues = _build_open_issue_state(current_issue_map, previous_open_issues, now_iso)
        next_resolved_issues = _build_resolved_issue_state(
            previous_resolved_issues,
            previous_open_issues,
            current_issue_map,
            resolved_issue_keys,
            now_iso,
        )
        next_resolved_issues = prune_resolved_issues(
            next_resolved_issues,
            get_notification_state_retention_days(),
        )

        new_issue_records = [current_issue_map[issue_key] for issue_key in new_issue_keys]
        resolved_issue_records = [
            next_resolved_issues[issue_key]
            for issue_key in resolved_issue_keys
            if issue_key in next_resolved_issues
        ]

        slack_text = _build_slack_summary_text(
            competitions,
            total_matches,
            total_new_matches,
            len(next_open_issues),
            new_issue_records,
            resolved_issue_records,
        )
        slack_sent = send_slack_message(slack_text)
        if slack_sent:
            logger.info("Slack summary sent to #notifications-fmm")
        else:
            logger.info("Slack not sent; will retry new issues on next run")

        # Always save notification state so deduplication works across runs
        notification_state_mgr.save_state(
            {
                "open_issues": next_open_issues,
                "resolved_issues": next_resolved_issues,
            }
        )

        return True
        
    except Exception as e:
        logger.error(f"Error creating Excel file: {str(e)}")
        return False


WHITELIST_HEADERS = ['Competition ID (required)', 'Name (auto-filled)', 'League ID (auto-filled)', 'League Name (auto-filled)', 'Start Date', 'End Date', 'Added Date']


def write_whitelist_sheet(wb, whitelist_config):
    """Add a 'Whitelist' tab to the workbook with the competition whitelist data."""
    if "Whitelist" in wb.sheetnames:
        del wb["Whitelist"]

    ws = wb.create_sheet("Whitelist")

    # Instructions row
    instruction_font = Font(italic=True, color="666666", size=10)
    instruction_cell = ws.cell(row=1, column=1, value="To add a competition: enter its ID in column A on a new row. Name, League ID, and League Name will be filled automatically on the next run.")
    instruction_cell.font = instruction_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(WHITELIST_HEADERS))

    # Header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col_idx, header in enumerate(WHITELIST_HEADERS, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    competitions = whitelist_config.get('active_competitions', [])
    for row_idx, comp in enumerate(competitions, 3):
        ws.cell(row=row_idx, column=1, value=comp.get('id', ''))
        ws.cell(row=row_idx, column=2, value=comp.get('name', ''))
        ws.cell(row=row_idx, column=3, value=comp.get('league_id', ''))
        ws.cell(row=row_idx, column=4, value=comp.get('league_name', ''))
        ws.cell(row=row_idx, column=5, value=comp.get('start_date', ''))
        ws.cell(row=row_idx, column=6, value=comp.get('end_date', ''))
        ws.cell(row=row_idx, column=7, value=comp.get('added_date', ''))

    # Auto-fit column widths
    for col_idx in range(1, len(WHITELIST_HEADERS) + 1):
        max_len = len(WHITELIST_HEADERS[col_idx - 1])
        for row in ws.iter_rows(min_row=3, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 4, 50)

    logger.info(f"Whitelist tab written with {len(competitions)} competitions")


def read_whitelist_from_excel(excel_path):
    """
    Read the competition whitelist from the 'Whitelist' tab of the Excel file.
    Returns a whitelist config dict compatible with competition_whitelist.json format,
    or None if the tab doesn't exist or can't be read.
    """
    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        if "Whitelist" not in wb.sheetnames:
            wb.close()
            return None

        ws = wb["Whitelist"]
        competitions = []
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row or not row[0]:
                continue
            comp_id = row[0]
            if isinstance(comp_id, str):
                try:
                    comp_id = int(comp_id)
                except ValueError:
                    continue
            elif not isinstance(comp_id, (int, float)):
                continue
            comp_id = int(comp_id)

            league_id = row[2] if len(row) > 2 else ''
            if isinstance(league_id, str):
                try:
                    league_id = int(league_id)
                except ValueError:
                    league_id = 0
            elif isinstance(league_id, (int, float)):
                league_id = int(league_id)
            else:
                league_id = 0

            comp = {
                'id': comp_id,
                'name': str(row[1]).strip() if len(row) > 1 and row[1] else '',
                'league_id': league_id,
                'league_name': str(row[3]).strip() if len(row) > 3 and row[3] else '',
            }
            if len(row) > 4 and row[4]:
                comp['start_date'] = str(row[4]).strip()
            if len(row) > 5 and row[5]:
                comp['end_date'] = str(row[5]).strip()
            if len(row) > 6 and row[6]:
                comp['added_date'] = str(row[6]).strip()

            competitions.append(comp)

        wb.close()

        if not competitions:
            return None

        logger.info(f"Read {len(competitions)} competitions from Excel whitelist tab")
        return {
            'active_competitions': competitions,
            'total_competitions': len(competitions),
        }
    except Exception as e:
        logger.error(f"Error reading whitelist from Excel: {e}")
        return None
